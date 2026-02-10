"""Command-line interface for pystylometry.

Usage:
    pystylometry-drift <file> [--window-size=N] [--stride=N] [--mode=MODE] [--json]
    pystylometry-drift <file> --plot [output.png]
    pystylometry-tokenize <file> [--json] [--metadata] [--stats]
    bnc --input-file <file> [--output-file <file>] [--format csv|html|json]

Example:
    pystylometry-drift manuscript.txt
    pystylometry-drift manuscript.txt --window-size=500 --stride=250
    pystylometry-drift manuscript.txt --json
    pystylometry-drift manuscript.txt --plot
    pystylometry-drift manuscript.txt --plot drift_report.png
    pystylometry-tokenize manuscript.txt
    pystylometry-tokenize manuscript.txt --json --metadata
    pystylometry-tokenize manuscript.txt --stats
    bnc --input-file manuscript.txt
    bnc --input-file manuscript.txt --output-file report.html --format html
    bnc -i manuscript.txt --format json
"""

from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path


def drift_cli() -> None:
    """CLI entry point for Kilgarriff drift detection."""
    parser = argparse.ArgumentParser(
        prog="pystylometry-drift",
        description="Detect stylistic drift within a document using Kilgarriff chi-squared.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  pystylometry-drift manuscript.txt
  pystylometry-drift manuscript.txt --window-size=500 --stride=250
  pystylometry-drift manuscript.txt --mode=all_pairs --json
  pystylometry-drift manuscript.txt --plot
  pystylometry-drift manuscript.txt --plot report.png
  pystylometry-drift manuscript.txt --plot timeline.png --plot-type=timeline
  pystylometry-drift manuscript.txt --jsx report.html --plot-type=report
  pystylometry-drift manuscript.txt --viz-all ./output  # All PNG + HTML

Pattern Signatures:
  consistent          Low, stable χ² across pairs (natural human writing)
  gradual_drift       Slowly increasing trend (author fatigue, topic shift)
  sudden_spike        One pair has high χ² (pasted content, different author)
  suspiciously_uniform Near-zero variance (possible AI generation)
""",
    )

    parser.add_argument(
        "file",
        type=Path,
        help="Path to text file to analyze",
    )
    parser.add_argument(
        "--window-size",
        type=int,
        default=1000,
        help="Number of tokens per window (default: 1000)",
    )
    parser.add_argument(
        "--stride",
        type=int,
        default=500,
        help="Tokens to advance between windows (default: 500)",
    )
    parser.add_argument(
        "--mode",
        choices=["sequential", "all_pairs", "fixed_lag"],
        default="sequential",
        help="Comparison mode (default: sequential)",
    )
    parser.add_argument(
        "--n-words",
        type=int,
        default=500,
        help="Most frequent words to analyze (default: 500)",
    )
    parser.add_argument(
        "--json",
        action="store_true",
        help="Output results as JSON",
    )
    parser.add_argument(
        "--plot",
        nargs="?",
        const="",
        default=None,
        metavar="OUTPUT",
        help="Generate visualization (optional: path to save, otherwise displays interactively)",
    )
    parser.add_argument(
        "--plot-type",
        choices=["report", "timeline"],
        default="report",
        help="Visualization type: report (multi-panel) or timeline (line chart)",
    )
    parser.add_argument(
        "--jsx",
        metavar="OUTPUT_FILE",
        help="Export interactive visualization as standalone HTML (uses --plot-type)",
    )
    parser.add_argument(
        "--viz-all",
        metavar="OUTPUT_DIR",
        type=Path,
        help="Generate ALL visualizations (PNG + HTML) to directory for testing",
    )

    args = parser.parse_args()

    # Validate file exists
    if not args.file.exists():
        print(f"Error: File not found: {args.file}", file=sys.stderr)
        sys.exit(1)

    # Read file
    try:
        text = args.file.read_text(encoding="utf-8")
    except Exception as e:
        print(f"Error reading file: {e}", file=sys.stderr)
        sys.exit(1)

    # Determine output mode
    if args.viz_all:
        output_mode = "All Visualizations (PNG + HTML)"
        output_dest = str(args.viz_all)
    elif args.jsx:
        output_mode = f"Interactive HTML ({args.plot_type})"
        output_dest = args.jsx
    elif args.plot is not None:
        output_mode = f"Plot ({args.plot_type})"
        output_dest = args.plot if args.plot else "interactive display"
    elif args.json:
        output_mode = "JSON"
        output_dest = "stdout"
    else:
        output_mode = "Text Report"
        output_dest = "stdout"

    # Calculate file stats
    token_count = len(text.split())
    char_count = len(text)

    # Print professional intro banner
    print()
    print("  PYSTYLOMETRY — Kilgarriff Chi-Squared Drift Detection")
    print("  ═══════════════════════════════════════════════════════════════════════")
    print()
    print("  INPUT")
    print("  ───────────────────────────────────────────────────────────────────────")
    print(f"    File:              {args.file}")
    print(f"    Size:              {char_count:,} characters / {token_count:,} tokens")
    print()
    print("  PARAMETERS")
    print("  ───────────────────────────────────────────────────────────────────────")
    print(f"    Window size:       {args.window_size} tokens")
    print(f"    Stride:            {args.stride} tokens")
    print(
        f"    Overlap:           {((args.window_size - args.stride) / args.window_size) * 100:.0f}%"
    )
    print(f"    Comparison mode:   {args.mode}")
    print(f"    Top N words:       {args.n_words}")
    print()
    print("  OUTPUT")
    print("  ───────────────────────────────────────────────────────────────────────")
    print(f"    Format:            {output_mode}")
    print(f"    Destination:       {output_dest}")
    print()
    print("  Running analysis...")
    print()

    # Import here to avoid slow startup
    from pystylometry.consistency import compute_kilgarriff_drift

    # Run analysis
    result = compute_kilgarriff_drift(
        text,
        window_size=args.window_size,
        stride=args.stride,
        comparison_mode=args.mode,
        n_words=args.n_words,
    )

    # Handle --viz-all: generate all visualizations for testing
    if args.viz_all:
        output_dir = args.viz_all
        output_dir.mkdir(parents=True, exist_ok=True)
        label = args.file.stem

        from pystylometry.viz.jsx import export_drift_timeline_jsx

        generated = []

        # Write chunks to subdirectory
        chunks_dir = output_dir / "chunks"
        chunks_dir.mkdir(parents=True, exist_ok=True)

        # Re-create windows to get chunk text (simple word-based chunking)
        words = text.split()
        chunk_texts = []
        start = 0
        chunk_idx = 0
        while start + args.window_size <= len(words):
            chunk_words = words[start : start + args.window_size]
            chunk_text = " ".join(chunk_words)
            chunk_texts.append(chunk_text)

            # Write chunk file
            chunk_path = chunks_dir / f"chunk_{chunk_idx:03d}.txt"
            chunk_path.write_text(chunk_text, encoding="utf-8")
            chunk_idx += 1
            start += args.stride

        print(f"  Created: {chunks_dir}/ ({len(chunk_texts)} chunks)")

        # Generate timeline HTML with chunk content
        out_path = output_dir / "drift-detection.html"
        export_drift_timeline_jsx(
            result,
            output_file=out_path,
            title=f"Drift Timeline: {label}",
            chunks=chunk_texts,
        )
        generated.append(out_path)
        print(f"  Created: {out_path}")

        print()
        n_viz, n_chunks = len(generated), len(chunk_texts)
        print(f"Generated {n_viz} visualizations + {n_chunks} chunks to: {output_dir.resolve()}")
        sys.exit(0)

    # Handle JSX export (generates standalone HTML)
    if args.jsx:
        from pystylometry.viz.jsx import (
            export_drift_report_jsx,
            export_drift_timeline_jsx,
        )

        label = args.file.stem

        if args.plot_type == "timeline":
            output_path = export_drift_timeline_jsx(
                result,
                output_file=args.jsx,
                title=f"Drift Timeline: {label}",
            )
        else:  # report (default)
            output_path = export_drift_report_jsx(
                result,
                output_file=args.jsx,
                label=label,
            )

        abs_path = output_path.resolve()
        file_url = f"file://{abs_path}"
        print(f"Interactive visualization saved to: {output_path}")
        print(f"Open in browser: {file_url}")
        sys.exit(0)

    # Handle plot output
    if args.plot is not None:
        try:
            from pystylometry.viz import plot_drift_report, plot_drift_timeline
        except ImportError:
            print(
                "Error: Visualization requires optional dependencies.",
                file=sys.stderr,
            )
            print(
                "Install with: pip install pystylometry[viz] or poetry install --with viz",
                file=sys.stderr,
            )
            sys.exit(1)

        plot_output: str | None = args.plot if args.plot else None
        label = args.file.stem

        if args.plot_type == "timeline":
            plot_drift_timeline(result, output=plot_output, title=f"Drift Timeline: {label}")
        else:  # report (default)
            plot_drift_report(result, label=label, output=plot_output)

        if plot_output:
            print(f"Visualization saved to: {plot_output}")
        sys.exit(0)

    if args.json:
        # JSON output
        output = {
            "status": result.status,
            "status_message": result.status_message,
            "pattern": result.pattern,
            "pattern_confidence": result.pattern_confidence,
            "mean_chi_squared": result.mean_chi_squared,
            "std_chi_squared": result.std_chi_squared,
            "max_chi_squared": result.max_chi_squared,
            "min_chi_squared": result.min_chi_squared,
            "max_location": result.max_location,
            "trend": result.trend,
            "window_size": result.window_size,
            "stride": result.stride,
            "overlap_ratio": result.overlap_ratio,
            "window_count": result.window_count,
            "comparison_mode": result.comparison_mode,
        }
        print(json.dumps(output, indent=2))
    else:
        # Human-readable output
        print("=" * 60)
        print("STYLISTIC DRIFT ANALYSIS")
        print("=" * 60)
        print(f"File: {args.file}")
        print(f"Status: {result.status}")
        print()

        if result.status == "insufficient_data":
            print(f"⚠️  {result.status_message}")
            print()
            print(f"Windows created: {result.window_count}")
            print("Minimum required: 3")
            print()
            print("Try reducing --window-size or --stride to create more windows.")
            sys.exit(0)

        print("PATTERN DETECTED")
        print("-" * 40)
        print(f"  Pattern: {result.pattern}")
        print(f"  Confidence: {result.pattern_confidence:.1%}")
        print()

        if result.pattern == "consistent":
            print("  ✓ Text shows consistent writing style throughout.")
        elif result.pattern == "gradual_drift":
            print("  ↗ Text shows gradual stylistic drift over its length.")
            print("    Possible causes: author fatigue, topic evolution, revision.")
        elif result.pattern == "sudden_spike":
            print("  ⚡ Text contains a sudden stylistic discontinuity.")
            loc = result.max_location
            print(f"    Location: Between windows {loc} and {loc + 1}")
            print("    Possible causes: pasted content, different author, major edit.")
        elif result.pattern == "suspiciously_uniform":
            print("  ⚠️  Text shows unusually uniform style (near-zero variance).")
            print("    Possible causes: AI-generated content, heavy editing, templated text.")

        print()
        print("CHI-SQUARED STATISTICS")
        print("-" * 40)
        print(f"  Mean χ²:  {result.mean_chi_squared:.2f}")
        print(f"  Std χ²:   {result.std_chi_squared:.2f}")
        print(f"  Min χ²:   {result.min_chi_squared:.2f}")
        print(f"  Max χ²:   {result.max_chi_squared:.2f}")
        print(f"  Trend:    {result.trend:+.4f}")
        print()

        print("WINDOW CONFIGURATION")
        print("-" * 40)
        print(f"  Window size:    {result.window_size} tokens")
        print(f"  Stride:         {result.stride} tokens")
        print(f"  Overlap:        {result.overlap_ratio:.1%}")
        print(f"  Windows:        {result.window_count}")
        print(f"  Comparisons:    {len(result.pairwise_scores)}")
        print()

        if result.status == "marginal_data":
            print(f"⚠️  {result.status_message}")
            print()


def viewer_cli() -> None:
    """CLI entry point for generating a standalone drift viewer."""
    parser = argparse.ArgumentParser(
        prog="pystylometry-viewer",
        description="Generate a standalone HTML drift analysis viewer.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
This generates a self-contained HTML file that users can open in any browser
to analyze their own text files. No Python or server required - just share
the HTML file and anyone can use it.

Examples:
  pystylometry-viewer drift_analyzer.html
  pystylometry-viewer ~/Desktop/analyzer.html --title "My Drift Analyzer"

The generated viewer includes:
  - Drag-and-drop file upload
  - Configurable analysis parameters
  - Interactive timeline visualization
  - Client-side Kilgarriff chi-squared implementation
""",
    )

    parser.add_argument(
        "output",
        type=Path,
        help="Path to write the HTML viewer file",
    )
    parser.add_argument(
        "--title",
        default="Stylistic Drift Analyzer",
        help="Page title (default: 'Stylistic Drift Analyzer')",
    )

    args = parser.parse_args()

    from pystylometry.viz.jsx import export_drift_viewer

    output_path = export_drift_viewer(args.output, title=args.title)

    abs_path = output_path.resolve()
    file_url = f"file://{abs_path}"

    print()
    print("  PYSTYLOMETRY — Standalone Drift Viewer")
    print("  ═══════════════════════════════════════════════════════════════════════")
    print()
    print(f"  Generated: {output_path}")
    print(f"  Open in browser: {file_url}")
    print()
    print("  This viewer can be shared with anyone. Users can:")
    print("    • Drag-and-drop or upload .txt files")
    print("    • Configure analysis parameters")
    print("    • View interactive drift timeline")
    print("    • Click points to see chunk comparisons")
    print()


def tokenize_cli() -> None:
    """CLI entry point for stylometric tokenization."""
    parser = argparse.ArgumentParser(
        prog="pystylometry-tokenize",
        description="Tokenize text for stylometric analysis.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  pystylometry-tokenize manuscript.txt
  pystylometry-tokenize manuscript.txt --json
  pystylometry-tokenize manuscript.txt --json --metadata
  pystylometry-tokenize manuscript.txt --stats
  pystylometry-tokenize manuscript.txt -U --expand-contractions
  pystylometry-tokenize manuscript.txt --min-length 3 --strip-numbers
""",
    )

    parser.add_argument(
        "file",
        type=Path,
        help="Path to text file to tokenize",
    )

    # Output mode
    output_group = parser.add_argument_group("output")
    output_group.add_argument(
        "-j",
        "--json",
        action="store_true",
        help="Output as JSON (list of strings, or list of objects with --metadata)",
    )
    output_group.add_argument(
        "-m",
        "--metadata",
        action="store_true",
        help="Include token type and position metadata (implies --json)",
    )
    output_group.add_argument(
        "-s",
        "--stats",
        action="store_true",
        help="Show tokenization statistics instead of tokens",
    )

    # Core behavior
    behavior_group = parser.add_argument_group("behavior")
    behavior_group.add_argument(
        "-U",
        "--no-lowercase",
        action="store_true",
        help="Preserve original case (default: lowercase)",
    )
    behavior_group.add_argument(
        "-e",
        "--expand-contractions",
        action="store_true",
        help="Expand contractions (it's -> it is)",
    )
    behavior_group.add_argument(
        "-n",
        "--strip-numbers",
        action="store_true",
        help="Remove numeric tokens",
    )
    behavior_group.add_argument(
        "--keep-punctuation",
        action="store_true",
        help="Keep punctuation tokens (default: stripped)",
    )

    # Filtering
    filter_group = parser.add_argument_group("filtering")
    filter_group.add_argument(
        "--min-length",
        type=int,
        default=1,
        metavar="N",
        help="Minimum token length (default: 1)",
    )
    filter_group.add_argument(
        "--max-length",
        type=int,
        default=None,
        metavar="N",
        help="Maximum token length (default: unlimited)",
    )
    filter_group.add_argument(
        "--preserve-urls",
        action="store_true",
        help="Keep URL tokens",
    )
    filter_group.add_argument(
        "--preserve-emails",
        action="store_true",
        help="Keep email tokens",
    )
    filter_group.add_argument(
        "--preserve-hashtags",
        action="store_true",
        help="Keep hashtag tokens",
    )
    filter_group.add_argument(
        "--preserve-mentions",
        action="store_true",
        help="Keep @mention tokens",
    )

    # Advanced
    advanced_group = parser.add_argument_group("advanced")
    advanced_group.add_argument(
        "--expand-abbreviations",
        action="store_true",
        help="Expand abbreviations (Dr. -> Doctor)",
    )
    advanced_group.add_argument(
        "--strip-accents",
        action="store_true",
        help="Remove accents from characters",
    )
    advanced_group.add_argument(
        "--no-clean",
        action="store_true",
        help="Skip text cleaning (italics, brackets, page markers)",
    )
    advanced_group.add_argument(
        "--no-unicode-normalize",
        action="store_true",
        help="Skip unicode normalization",
    )

    args = parser.parse_args()

    # --- ANSI colors ---
    use_color = sys.stderr.isatty()

    def _c(code: str, text: str) -> str:
        return f"\033[{code}m{text}\033[0m" if use_color else text

    bold = lambda t: _c("1", t)  # noqa: E731
    dim = lambda t: _c("2", t)  # noqa: E731
    cyan = lambda t: _c("36", t)  # noqa: E731
    green = lambda t: _c("32", t)  # noqa: E731
    yellow = lambda t: _c("33", t)  # noqa: E731

    # --- Validate file ---
    if not args.file.exists():
        print(f"Error: File not found: {args.file}", file=sys.stderr)
        sys.exit(1)

    try:
        text = args.file.read_text(encoding="utf-8")
    except Exception as e:
        print(f"Error reading file: {e}", file=sys.stderr)
        sys.exit(1)

    # --- Build Tokenizer kwargs ---
    tokenizer_kwargs = {
        "lowercase": not args.no_lowercase,
        "min_length": args.min_length,
        "max_length": args.max_length,
        "strip_numbers": args.strip_numbers,
        "strip_punctuation": not args.keep_punctuation,
        "preserve_urls": args.preserve_urls,
        "preserve_emails": args.preserve_emails,
        "preserve_hashtags": args.preserve_hashtags,
        "preserve_mentions": args.preserve_mentions,
        "expand_contractions": args.expand_contractions,
        "expand_abbreviations": args.expand_abbreviations,
        "strip_accents": args.strip_accents,
        "normalize_unicode": not args.no_unicode_normalize,
        "clean_text": not args.no_clean,
    }

    # Collect active options for banner
    active_opts = []
    if args.no_lowercase:
        active_opts.append("preserve case")
    if args.expand_contractions:
        active_opts.append("expand contractions")
    if args.expand_abbreviations:
        active_opts.append("expand abbreviations")
    if args.strip_numbers:
        active_opts.append("strip numbers")
    if args.keep_punctuation:
        active_opts.append("keep punctuation")
    if args.strip_accents:
        active_opts.append("strip accents")
    if args.no_clean:
        active_opts.append("skip cleaning")
    if args.no_unicode_normalize:
        active_opts.append("skip unicode normalization")
    if args.preserve_urls:
        active_opts.append("preserve URLs")
    if args.preserve_emails:
        active_opts.append("preserve emails")
    if args.preserve_hashtags:
        active_opts.append("preserve hashtags")
    if args.preserve_mentions:
        active_opts.append("preserve mentions")
    if args.min_length > 1:
        active_opts.append(f"min length {args.min_length}")
    if args.max_length is not None:
        active_opts.append(f"max length {args.max_length}")

    # Determine output format
    if args.stats:
        output_format = "Statistics"
    elif args.metadata:
        output_format = "JSON (with metadata)"
    elif args.json:
        output_format = "JSON"
    else:
        output_format = "One token per line"

    # --- Banner (to stderr so stdout stays pipeable) ---
    char_count = len(text)
    line_count = text.count("\n") + 1

    banner = sys.stderr
    print(file=banner)
    print(f"  {bold('PYSTYLOMETRY')} {dim('—')} {cyan('Stylometric Tokenizer')}", file=banner)
    print(f"  {dim('═' * 71)}", file=banner)
    print(file=banner)
    print(f"  {bold('INPUT')}", file=banner)
    print(f"  {dim('─' * 71)}", file=banner)
    print(f"    File:              {args.file}", file=banner)
    print(f"    Size:              {char_count:,} characters / {line_count:,} lines", file=banner)
    print(file=banner)
    print(f"  {bold('CONFIGURATION')}", file=banner)
    print(f"  {dim('─' * 71)}", file=banner)
    print(f"    Case:              {'preserve' if args.no_lowercase else 'lowercase'}", file=banner)
    print(
        f"    Punctuation:       {'keep' if args.keep_punctuation else 'strip'}",
        file=banner,
    )
    print(
        f"    Contractions:      {'expand' if args.expand_contractions else 'preserve'}",
        file=banner,
    )
    print(f"    Numbers:           {'strip' if args.strip_numbers else 'keep'}", file=banner)
    if active_opts:
        print(f"    Active options:    {', '.join(active_opts)}", file=banner)
    print(file=banner)
    print(f"  {bold('OUTPUT')}", file=banner)
    print(f"  {dim('─' * 71)}", file=banner)
    print(f"    Format:            {output_format}", file=banner)
    print(file=banner)

    # --- Tokenize ---
    from pystylometry.tokenizer import Tokenizer

    tokenizer = Tokenizer(**tokenizer_kwargs)

    if args.stats:
        stats = tokenizer.get_statistics(text)
        print(f"  {bold('RESULTS')}", file=banner)
        print(f"  {dim('─' * 71)}", file=banner)
        print(f"    Total tokens:      {green(f'{stats.total_tokens:,}')}", file=banner)
        print(f"    Unique tokens:     {green(f'{stats.unique_tokens:,}')}", file=banner)
        print(f"    Word tokens:       {stats.word_tokens:,}", file=banner)
        print(f"    Number tokens:     {stats.number_tokens:,}", file=banner)
        print(f"    Punctuation:       {stats.punctuation_tokens:,}", file=banner)
        print(f"    URLs:              {stats.url_tokens:,}", file=banner)
        print(f"    Emails:            {stats.email_tokens:,}", file=banner)
        print(f"    Hashtags:          {stats.hashtag_tokens:,}", file=banner)
        print(f"    Mentions:          {stats.mention_tokens:,}", file=banner)
        print(f"    Avg length:        {stats.average_token_length:.1f}", file=banner)
        print(f"    Min length:        {stats.min_token_length}", file=banner)
        print(f"    Max length:        {stats.max_token_length}", file=banner)
        print(file=banner)

        if args.json:
            import dataclasses

            print(json.dumps(dataclasses.asdict(stats), indent=2))

    elif args.metadata or (args.json and args.metadata):
        metadata_list = tokenizer.tokenize_with_metadata(text)
        count = len(metadata_list)
        print(
            f"  {yellow('Tokenizing...')} {green(f'{count:,}')} tokens extracted",
            file=banner,
        )
        print(file=banner)
        output = [
            {
                "token": m.token,
                "start": m.start,
                "end": m.end,
                "type": m.token_type,
            }
            for m in metadata_list
        ]
        print(json.dumps(output, indent=2))

    elif args.json:
        tokens = tokenizer.tokenize(text)
        count = len(tokens)
        print(
            f"  {yellow('Tokenizing...')} {green(f'{count:,}')} tokens extracted",
            file=banner,
        )
        print(file=banner)
        print(json.dumps(tokens, indent=2))

    else:
        tokens = tokenizer.tokenize(text)
        count = len(tokens)
        print(
            f"  {yellow('Tokenizing...')} {green(f'{count:,}')} tokens extracted",
            file=banner,
        )
        print(file=banner)
        for token in tokens:
            print(token)


def bnc_frequency_cli() -> None:
    """CLI entry point for BNC word frequency analysis."""
    parser = argparse.ArgumentParser(
        prog="bnc",
        description="Analyze word frequencies against the British National Corpus (BNC).",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  bnc --input-file manuscript.txt
  bnc --input-file manuscript.txt --output-file report.html
  bnc --input-file manuscript.txt --format json
  bnc --input-file manuscript.txt --overuse-threshold 2.0 --min-mentions 3
  bnc --input-file manuscript.txt --no-wordnet
  bnc --input-dir ~/ebooks/Author --format excel --output-file author.xlsx

Output:
  Generates a report with three sections:
  - Not in BNC: Words not found in the corpus (with WordNet status, character type)
  - Most Underused: Words appearing less frequently than expected
  - Most Overused: Words appearing more frequently than expected

Thresholds:
  Words with ratio > overuse-threshold are "overused"
  Words with ratio < underuse-threshold are "underused"
  Ratio = observed_count / expected_count (based on BNC frequencies)
""",
    )

    input_group = parser.add_mutually_exclusive_group(required=True)
    input_group.add_argument(
        "--input-file",
        "-i",
        type=Path,
        metavar="FILE",
        help="Path to text file to analyze",
    )
    input_group.add_argument(
        "--input-dir",
        "-d",
        type=Path,
        metavar="DIR",
        help="Directory of .txt files to analyze as a single corpus (requires --output-file)",
    )
    parser.add_argument(
        "--output-file",
        "-o",
        type=Path,
        default=None,
        metavar="FILE",
        help="Output file (default: <input>_bnc_frequency.<ext> based on --format)",
    )
    parser.add_argument(
        "--overuse-threshold",
        type=float,
        default=1.3,
        metavar="N",
        help="Ratio above which words are considered overused (default: 1.3)",
    )
    parser.add_argument(
        "--underuse-threshold",
        type=float,
        default=0.8,
        metavar="N",
        help="Ratio below which words are considered underused (default: 0.8)",
    )
    parser.add_argument(
        "--min-mentions",
        type=int,
        default=1,
        metavar="N",
        help="Minimum word occurrences to include (default: 1)",
    )
    parser.add_argument(
        "--no-wordnet",
        action="store_true",
        help="Skip WordNet lookup for unknown words",
    )
    parser.add_argument(
        "--format",
        choices=["csv", "html", "json", "excel"],
        default="csv",
        help="Output format: csv (tab-delimited), html (interactive), json, excel (default: csv)",
    )
    parser.add_argument(
        "--combinatoric",
        action="store_true",
        help="Pairwise comparison of all files in --input-dir (requires --input-dir)",
    )

    args = parser.parse_args()

    # Import rich for colored output
    from rich.console import Console
    from rich.panel import Panel
    from rich.table import Table
    from rich.text import Text

    console = Console(stderr=True)

    # ── Validation ──
    if args.combinatoric and not args.input_dir:
        console.print("[red]Error:[/red] --combinatoric requires --input-dir")
        sys.exit(1)

    # ── Input loading ──
    file_count: int = 1
    txt_files: list[Path] = []

    if args.input_dir:
        if not args.input_dir.is_dir():
            console.print(f"[red]Error:[/red] Directory not found: {args.input_dir}")
            sys.exit(1)
        if not args.output_file:
            console.print("[red]Error:[/red] --output-file is required when using --input-dir")
            sys.exit(1)

        txt_files = sorted(args.input_dir.glob("*.txt"))
        if not txt_files:
            console.print(f"[red]Error:[/red] No .txt files found in: {args.input_dir}")
            sys.exit(1)

        # Read all files (used for both standard and combinatoric modes)
        parts: list[str] = []
        for tf in txt_files:
            try:
                parts.append(tf.read_text(encoding="utf-8"))
            except Exception as e:
                console.print(f"[red]Error reading {tf.name}:[/red] {e}")
                sys.exit(1)
        text = "\n".join(parts)
        file_count = len(txt_files)
    else:
        # --input-file mode (existing behavior)
        if not args.input_file.exists():
            console.print(f"[red]Error:[/red] File not found: {args.input_file}")
            sys.exit(1)
        try:
            text = args.input_file.read_text(encoding="utf-8")
        except Exception as e:
            console.print(f"[red]Error reading file:[/red] {e}")
            sys.exit(1)

    # Determine output path (extension based on format)
    suffix_map = {"csv": ".tsv", "html": ".html", "json": ".json", "excel": ".xlsx"}
    if args.output_file:
        output_path = args.output_file
    else:
        suffix = suffix_map[args.format]
        output_path = args.input_file.with_name(f"{args.input_file.stem}_bnc_frequency{suffix}")

    # Calculate file stats
    token_count = len(text.split())
    char_count = len(text)

    # Print header
    console.print()
    header = Text()
    header.append("PYSTYLOMETRY", style="bold cyan")
    header.append(" — ", style="dim")
    header.append("BNC Word Frequency Analysis", style="bold white")
    console.print(Panel(header, border_style="cyan"))

    # Input section
    console.print()
    console.print("[bold]INPUT[/bold]", style="cyan")
    console.print("─" * 60, style="dim")
    if args.input_dir:
        console.print(f"  Dir:     [white]{args.input_dir}[/white]")
        console.print(f"  Files:   [green]{file_count}[/green] .txt files")
    else:
        console.print(f"  File:    [white]{args.input_file}[/white]")
    _sz = f"[green]{char_count:,}[/green] chars / [green]{token_count:,}[/green] tokens"
    console.print(f"  Size:    {_sz}")

    # Probe optional dependency availability for banner
    from importlib.metadata import PackageNotFoundError
    from importlib.metadata import version as _pkg_ver

    try:
        _bnc_ver: str = _pkg_ver("bnc-lookup")
    except PackageNotFoundError:
        _bnc_ver = "[red]not installed[/red]"

    try:
        _gn_ver: str = _pkg_ver("gngram-lookup")
    except PackageNotFoundError:
        _gn_ver = "[dim]not installed[/dim]"

    try:
        _wn_ver: str = _pkg_ver("wordnet-lookup")
    except PackageNotFoundError:
        _wn_ver = "[dim]not installed[/dim]"

    # Parameters section
    console.print()
    console.print("[bold]PARAMETERS[/bold]", style="cyan")
    console.print("─" * 60, style="dim")
    console.print(f"  Overuse threshold:   [yellow]{args.overuse_threshold}x[/yellow]")
    console.print(f"  Underuse threshold:  [yellow]{args.underuse_threshold}x[/yellow]")
    console.print(f"  Min mentions:        [yellow]{args.min_mentions}[/yellow]")
    console.print(f"  BNC lookup:          [yellow]{_bnc_ver}[/yellow]")
    _wn_display = "disabled" if args.no_wordnet else _wn_ver
    console.print(f"  WordNet lookup:      [yellow]{_wn_display}[/yellow]")
    console.print(f"  Ngram lookup:        [yellow]{_gn_ver}[/yellow]")
    if args.combinatoric:
        console.print(f"  Combinatoric:        [yellow]yes ({file_count} files)[/yellow]")

    # Output section
    console.print()
    console.print("[bold]OUTPUT[/bold]", style="cyan")
    console.print("─" * 60, style="dim")
    fmt_display = {
        "csv": "Tab-delimited CSV",
        "html": "Interactive HTML",
        "json": "JSON",
        "excel": "Excel (.xlsx)",
    }
    console.print(f"  Format:      [magenta]{fmt_display[args.format]}[/magenta]")
    console.print(f"  Destination: [white]{output_path}[/white]")

    # Run analysis with spinner
    console.print()
    with console.status("[bold cyan]Running analysis...[/bold cyan]", spinner="dots"):
        from pystylometry.lexical.bnc_frequency import compute_bnc_frequency

        result = compute_bnc_frequency(
            text,
            overuse_threshold=args.overuse_threshold,
            underuse_threshold=args.underuse_threshold,
            include_wordnet=not args.no_wordnet,
            min_mentions=args.min_mentions,
        )

    # ── Combinatoric: per-file works counts ──
    works_overused: dict[str, int] = {}
    works_underused: dict[str, int] = {}
    works_notbnc: dict[str, int] = {}
    works_total: int = 0

    if args.combinatoric:
        works_total = len(txt_files)
        console.print()
        for tf in txt_files:
            console.print(
                f"  [dim]Scanning[/dim] [white]{tf.name}[/white]...",
                end="",
            )
            file_text = tf.read_text(encoding="utf-8")
            fr = compute_bnc_frequency(
                file_text,
                overuse_threshold=args.overuse_threshold,
                underuse_threshold=args.underuse_threshold,
                include_wordnet=not args.no_wordnet,
                min_mentions=args.min_mentions,
            )
            for w in fr.overused:
                works_overused[w.word] = works_overused.get(w.word, 0) + 1
            for w in fr.underused:
                works_underused[w.word] = works_underused.get(w.word, 0) + 1
            for w in fr.not_in_bnc:
                works_notbnc[w.word] = works_notbnc.get(w.word, 0) + 1
            console.print(" [green]✓[/green]")

    # Output results
    if args.format == "json":
        # classify_word provides the three-layer morphological taxonomy label
        # that replaces the five boolean flag columns.
        # See: https://github.com/craigtrim/pystylometry/issues/53
        from pystylometry.lexical.word_class import classify_word as _classify_json

        def _json_word(w_obj: object, cat_counts: dict[str, int]) -> dict[str, object]:
            """Build a JSON-serializable dict for a WordAnalysis."""
            from pystylometry.lexical.bnc_frequency import WordAnalysis

            assert isinstance(w_obj, WordAnalysis)
            d: dict[str, object] = {
                "word": w_obj.word,
                "observed": w_obj.observed,
                "expected": w_obj.expected,
            }
            if cat_counts:
                d["works"] = f"{cat_counts.get(w_obj.word, 0)}/{works_total}"
            d["ratio"] = w_obj.ratio
            d["in_wordnet"] = w_obj.in_wordnet
            d["in_gngram"] = w_obj.in_gngram
            d["classification"] = _classify_json(w_obj.word).label
            return d

        stats: dict[str, object] = {
            "total_tokens": result.total_tokens,
            "unique_tokens": result.unique_tokens,
            "overused_count": len(result.overused),
            "underused_count": len(result.underused),
            "not_in_bnc_count": len(result.not_in_bnc),
        }
        if works_total:
            stats["works_total"] = works_total

        def _sort_key(w_obj: object, cat_counts: dict[str, int]) -> tuple[int, float]:
            from pystylometry.lexical.bnc_frequency import WordAnalysis

            assert isinstance(w_obj, WordAnalysis)
            return (cat_counts.get(w_obj.word, 0), w_obj.ratio or 0)

        _over_sorted = (
            sorted(result.overused, key=lambda x: _sort_key(x, works_overused), reverse=True)
            if works_total
            else result.overused
        )
        _under_sorted = (
            sorted(result.underused, key=lambda x: _sort_key(x, works_underused), reverse=True)
            if works_total
            else result.underused
        )
        _notbnc_sorted = (
            sorted(result.not_in_bnc, key=lambda x: _sort_key(x, works_notbnc), reverse=True)
            if works_total
            else result.not_in_bnc
        )

        output = {
            "stats": stats,
            "overused": [_json_word(w, works_overused) for w in _over_sorted],
            "underused": [_json_word(w, works_underused) for w in _under_sorted],
            "not_in_bnc": [_json_word(w, works_notbnc) for w in _notbnc_sorted],
        }
        output_path.write_text(json.dumps(output, indent=2))
        console.print(f'[green]✓[/green] JSON saved to: [white]"{output_path}"[/white]')

    elif args.format == "csv":
        # classify_word provides the three-layer morphological taxonomy label
        # that replaces the five boolean flag columns.
        # See: https://github.com/craigtrim/pystylometry/issues/53
        from pystylometry.lexical.word_class import classify_word as _classify_csv

        # Tab-delimited output with category column
        if works_total:
            _hdr = (
                "category\tword\tobserved\texpected\tworks\tratio\tin_wordnet\tin_gngram"
                "\tclassification"
            )
        else:
            _hdr = (
                "category\tword\tobserved\texpected\tratio\tin_wordnet\tin_gngram"
                "\tclassification"
            )
        lines = [_hdr]

        def fmt_wordnet(val: bool | None) -> str:
            if val is True:
                return "yes"
            elif val is False:
                return "no"
            return ""

        def fmt_gngram(val: bool | None) -> str:
            if val is True:
                return "yes"
            elif val is False:
                return "no"
            return ""

        _csv_over = (
            sorted(
                result.overused,
                key=lambda x: (works_overused.get(x.word, 0), x.ratio or 0),
                reverse=True,
            )
            if works_total
            else result.overused
        )
        for w in _csv_over:
            expected = f"{w.expected:.2f}" if w.expected else ""
            ratio = f"{w.ratio:.4f}" if w.ratio else ""
            in_wn = fmt_wordnet(w.in_wordnet)
            in_gn = fmt_gngram(w.in_gngram)
            cls = _classify_csv(w.word).label
            if works_total:
                wk = f"{works_overused.get(w.word, 0)}/{works_total}"
                line = (
                    f"overused\t{w.word}\t{w.observed}\t{expected}\t{wk}\t{ratio}"
                    f"\t{in_wn}\t{in_gn}\t{cls}"
                )
            else:
                line = (
                    f"overused\t{w.word}\t{w.observed}\t{expected}\t{ratio}"
                    f"\t{in_wn}\t{in_gn}\t{cls}"
                )
            lines.append(line)

        _csv_under = (
            sorted(
                result.underused,
                key=lambda x: (works_underused.get(x.word, 0), x.ratio or 0),
                reverse=True,
            )
            if works_total
            else result.underused
        )
        for w in _csv_under:
            expected = f"{w.expected:.2f}" if w.expected else ""
            ratio = f"{w.ratio:.4f}" if w.ratio else ""
            in_wn = fmt_wordnet(w.in_wordnet)
            in_gn = fmt_gngram(w.in_gngram)
            cls = _classify_csv(w.word).label
            if works_total:
                wk = f"{works_underused.get(w.word, 0)}/{works_total}"
                line = (
                    f"underused\t{w.word}\t{w.observed}\t{expected}\t{wk}\t{ratio}"
                    f"\t{in_wn}\t{in_gn}\t{cls}"
                )
            else:
                line = (
                    f"underused\t{w.word}\t{w.observed}\t{expected}\t{ratio}"
                    f"\t{in_wn}\t{in_gn}\t{cls}"
                )
            lines.append(line)

        _csv_notbnc = (
            sorted(
                result.not_in_bnc,
                key=lambda x: (works_notbnc.get(x.word, 0), x.observed or 0),
                reverse=True,
            )
            if works_total
            else result.not_in_bnc
        )
        for w in _csv_notbnc:
            in_wn = fmt_wordnet(w.in_wordnet)
            in_gn = fmt_gngram(w.in_gngram)
            cls = _classify_csv(w.word).label
            if works_total:
                wk = f"{works_notbnc.get(w.word, 0)}/{works_total}"
                line = f"not-in-bnc\t{w.word}\t{w.observed}\t\t{wk}\t" f"\t{in_wn}\t{in_gn}\t{cls}"
            else:
                line = f"not-in-bnc\t{w.word}\t{w.observed}\t\t\t{in_wn}\t{in_gn}" f"\t{cls}"
            lines.append(line)

        output_path.write_text("\n".join(lines))
        console.print(f'[green]✓[/green] TSV saved to: [white]"{output_path}"[/white]')

    elif args.format == "excel":
        try:
            from openpyxl import Workbook  # type: ignore[import-untyped]
            from openpyxl.styles import (  # type: ignore[import-untyped]
                Alignment,
                Font,
                PatternFill,
            )
        except ImportError:
            console.print("[red]Error:[/red] Excel export requires openpyxl.")
            console.print("  Install with: [yellow]pip install pystylometry[excel][/yellow]")
            console.print("  Or for pipx: [yellow]pipx inject pystylometry openpyxl[/yellow]")
            sys.exit(1)

        wb = Workbook()

        # Remove default sheet
        wb.remove(wb.active)

        # Cell styles
        align = Alignment(horizontal="center", vertical="center")
        font_header = Font(bold=True, size=14)
        fill_header = PatternFill(
            start_color="D9E2F3", end_color="D9E2F3", fill_type="solid"
        )  # Soft blue-gray

        # ── Summary sheet (first tab) ──
        ws_summary = wb.create_sheet("summary")

        # ── Corpus section — books analysed ──
        book_names: list[str] = []
        if txt_files:
            book_names = [tf.stem for tf in txt_files]
        elif args.input_file:
            book_names = [args.input_file.stem]

        c = ws_summary.cell(row=1, column=1, value="Corpus")
        c.font = font_header
        c.fill = fill_header
        c = ws_summary.cell(row=1, column=2, value="")
        c.fill = fill_header
        c = ws_summary.cell(row=1, column=3, value="")
        c.fill = fill_header

        ws_summary.cell(row=2, column=1, value="Total Books")
        ws_summary.cell(row=2, column=2, value=len(book_names))
        for bi, bname in enumerate(book_names):
            ws_summary.cell(row=3 + bi, column=1, value=bname)

        # Offset: corpus header (1) + total row (1) + book rows + blank row
        top10_start = 3 + len(book_names) + 1

        # Headers: 3 columns side by side
        headers = ["Top 10 Overused", "Top 10 Underused", "Top 10 Not in BNC"]
        for ci, hdr in enumerate(headers, start=1):
            c = ws_summary.cell(row=top10_start, column=ci, value=hdr)
            c.font = font_header
            c.fill = fill_header

        # Top 10 words from each category (words only, no numbers)
        top_overused = [w.word for w in result.overused[:10]]
        top_underused = [w.word for w in result.underused[:10]]
        top_not_in_bnc = [w.word for w in result.not_in_bnc[:10]]

        for i in range(10):
            row = top10_start + 1 + i
            if i < len(top_overused):
                ws_summary.cell(row=row, column=1, value=top_overused[i])
            if i < len(top_underused):
                ws_summary.cell(row=row, column=2, value=top_underused[i])
            if i < len(top_not_in_bnc):
                ws_summary.cell(row=row, column=3, value=top_not_in_bnc[i])

        # ── Descriptive stats below the top-10 table ──
        stats_start = top10_start + 12  # 10 word rows + 1 blank
        stat_headers = ["Metric", "Count", "% of Total Tokens"]
        for ci, hdr in enumerate(stat_headers, start=1):
            c = ws_summary.cell(row=stats_start, column=ci, value=hdr)
            c.font = font_header
            c.fill = fill_header

        total = result.total_tokens or 1  # avoid division by zero
        stats_rows = [
            ("Total Tokens", f"{result.total_tokens:,}", ""),
            (
                "Unique Tokens",
                f"{result.unique_tokens:,}",
                f"{result.unique_tokens / total * 100:.2f}%",
            ),
            ("Overused", f"{len(result.overused):,}", f"{len(result.overused) / total * 100:.2f}%"),
            (
                "Underused",
                f"{len(result.underused):,}",
                f"{len(result.underused) / total * 100:.2f}%",
            ),
            (
                "Not in BNC",
                f"{len(result.not_in_bnc):,}",
                f"{len(result.not_in_bnc) / total * 100:.2f}%",
            ),
        ]
        for ri, (label, count, pct) in enumerate(stats_rows, start=stats_start + 1):
            ws_summary.cell(row=ri, column=1, value=label)
            ws_summary.cell(row=ri, column=2, value=count)
            ws_summary.cell(row=ri, column=3, value=pct)

        # Format summary sheet
        for col_letter in ("A", "B", "C"):
            ws_summary.column_dimensions[col_letter].width = 25
        for r in ws_summary.iter_rows():
            for cell in r:
                cell.alignment = align

        def fmt_wordnet_excel(val: bool | None) -> str:
            if val is True:
                return "true"
            elif val is False:
                return "false"
            return ""

        def fmt_gngram_excel(val: bool | None) -> str:
            if val is True:
                return "true"
            elif val is False:
                return "false"
            return ""

        # classify_word provides the three-layer morphological taxonomy label
        # that replaces the five boolean flag columns.
        # See: https://github.com/craigtrim/pystylometry/issues/53
        from pystylometry.lexical.word_class import classify_word as _classify_xl

        # Overused sheet (sorted by works desc, ratio desc when combinatoric)
        ws_over = wb.create_sheet("overused")
        if works_total:
            _over_hdr = [
                "word",
                "observed",
                "expected",
                "works",
                "ratio",
                "in_wordnet",
                "in_gngram",
                "classification",
            ]
        else:
            _over_hdr = [
                "word",
                "observed",
                "expected",
                "ratio",
                "in_wordnet",
                "in_gngram",
                "classification",
            ]
        ws_over.append(_over_hdr)
        _xl_over = (
            sorted(
                result.overused,
                key=lambda x: (works_overused.get(x.word, 0), x.ratio or 0),
                reverse=True,
            )
            if works_total
            else sorted(result.overused, key=lambda x: x.ratio or 0, reverse=True)
        )
        for w in _xl_over:
            in_wn = fmt_wordnet_excel(w.in_wordnet)
            in_gn = fmt_gngram_excel(w.in_gngram)
            cls = _classify_xl(w.word).label
            if works_total:
                row_data = [
                    w.word,
                    w.observed,
                    w.expected,
                    f"{works_overused.get(w.word, 0)}/{works_total}",
                    w.ratio,
                    in_wn,
                    in_gn,
                    cls,
                ]
            else:
                row_data = [
                    w.word,
                    w.observed,
                    w.expected,
                    w.ratio,
                    in_wn,
                    in_gn,
                    cls,
                ]
            ws_over.append(row_data)

        # Underused sheet (sorted by works desc, ratio desc when combinatoric)
        ws_under = wb.create_sheet("underused")
        if works_total:
            _under_hdr = [
                "word",
                "observed",
                "expected",
                "works",
                "ratio",
                "in_wordnet",
                "in_gngram",
                "classification",
            ]
        else:
            _under_hdr = [
                "word",
                "observed",
                "expected",
                "ratio",
                "in_wordnet",
                "in_gngram",
                "classification",
            ]
        ws_under.append(_under_hdr)
        _xl_under = (
            sorted(
                result.underused,
                key=lambda x: (works_underused.get(x.word, 0), x.ratio or 0),
                reverse=True,
            )
            if works_total
            else sorted(result.underused, key=lambda x: x.ratio or 0, reverse=True)
        )
        for w in _xl_under:
            in_wn = fmt_wordnet_excel(w.in_wordnet)
            in_gn = fmt_gngram_excel(w.in_gngram)
            cls = _classify_xl(w.word).label
            if works_total:
                row_data = [
                    w.word,
                    w.observed,
                    w.expected,
                    f"{works_underused.get(w.word, 0)}/{works_total}",
                    w.ratio,
                    in_wn,
                    in_gn,
                    cls,
                ]
            else:
                row_data = [
                    w.word,
                    w.observed,
                    w.expected,
                    w.ratio,
                    in_wn,
                    in_gn,
                    cls,
                ]
            ws_under.append(row_data)

        # Not in BNC sheet
        ws_notbnc = wb.create_sheet("not-in-bnc")
        if works_total:
            _notbnc_hdr = [
                "word",
                "observed",
                "works",
                "in_wordnet",
                "in_gngram",
                "classification",
            ]
        else:
            _notbnc_hdr = [
                "word",
                "observed",
                "in_wordnet",
                "in_gngram",
                "classification",
            ]
        ws_notbnc.append(_notbnc_hdr)

        _xl_notbnc = (
            sorted(
                result.not_in_bnc,
                key=lambda x: (works_notbnc.get(x.word, 0), x.observed or 0),
                reverse=True,
            )
            if works_total
            else result.not_in_bnc
        )
        for w in _xl_notbnc:
            in_wn = fmt_wordnet_excel(w.in_wordnet)
            in_gn = fmt_gngram_excel(w.in_gngram)
            cls = _classify_xl(w.word).label
            if works_total:
                row_data = [
                    w.word,
                    w.observed,
                    f"{works_notbnc.get(w.word, 0)}/{works_total}",
                    in_wn,
                    in_gn,
                    cls,
                ]
            else:
                row_data = [
                    w.word,
                    w.observed,
                    in_wn,
                    in_gn,
                    cls,
                ]
            ws_notbnc.append(row_data)

        # Apply header styling to data sheets (match summary headers)
        for ws in [ws_over, ws_under, ws_notbnc]:
            for cell in ws[1]:
                cell.font = font_header
                cell.fill = fill_header

        # Apply formatting to all sheets
        for ws in [ws_over, ws_under, ws_notbnc]:
            for col in ws.columns:
                col_letter = col[0].column_letter
                # Word column (A) and classification (last col) get 30, others 15
                ws.column_dimensions[col_letter].width = 30 if col_letter == "A" else 15
            for row in ws.iter_rows():
                for cell in row:
                    cell.alignment = align

        # Classification column is wider to accommodate dot-separated labels
        # like "apostrophe.contraction.negative"
        # Overused/underused: col H (with works) or G (without)
        _cls_col_freq = "H" if works_total else "G"
        for ws in [ws_over, ws_under]:
            ws.column_dimensions[_cls_col_freq].width = 35
        # Not-in-BNC: col F (with works) or E (without)
        _cls_col_notbnc = "F" if works_total else "E"
        ws_notbnc.column_dimensions[_cls_col_notbnc].width = 35

        # Apply number formatting to expected and ratio columns
        # With works: C=expected, E=ratio; without works: C=expected, D=ratio
        _ratio_col = "E" if works_total else "D"
        for ws in [ws_over, ws_under]:
            for row in range(2, ws.max_row + 1):  # Skip header row
                ws[f"C{row}"].number_format = "0.00"
                ws[f"{_ratio_col}{row}"].number_format = "0.00"

        # Boolean flag colors (true/false) for in_wordnet and in_gngram columns.
        # The old five boolean columns (unicode, numeric, apostrophe, hyphen,
        # other) have been replaced by the classification string column (#53).
        fill_flag_true = PatternFill(
            start_color="BDD7EE", end_color="BDD7EE", fill_type="solid"
        )  # Light blue
        fill_flag_false = PatternFill(
            start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"
        )  # Light peach

        # Boolean columns: in_wordnet and in_gngram only
        # Overused/underused: with works F,G; without E,F
        _over_flags = ("F", "G") if works_total else ("E", "F")
        for ws in [ws_over, ws_under]:
            for row in range(2, ws.max_row + 1):
                for col_letter in _over_flags:
                    cell = ws[f"{col_letter}{row}"]
                    if cell.value == "true":
                        cell.fill = fill_flag_true
                    elif cell.value == "false":
                        cell.fill = fill_flag_false

        # Not-in-bnc: with works D,E; without C,D
        _notbnc_flags = ("D", "E") if works_total else ("C", "D")
        for row in range(2, ws_notbnc.max_row + 1):
            for col_letter in _notbnc_flags:
                cell = ws_notbnc[f"{col_letter}{row}"]
                if cell.value == "true":
                    cell.fill = fill_flag_true
                elif cell.value == "false":
                    cell.fill = fill_flag_false

        # Row heights: 25 for headers, 20 for data rows — all sheets
        for ws in [ws_summary, ws_over, ws_under, ws_notbnc]:
            for row_num in range(1, ws.max_row + 1):
                ws.row_dimensions[row_num].height = 20
            # Header rows get 25
            ws.row_dimensions[1].height = 25
        # Summary has additional header rows for top-10 and stats sections
        ws_summary.row_dimensions[top10_start].height = 25
        ws_summary.row_dimensions[stats_start].height = 25

        wb.save(output_path)
        console.print(f'[green]✓[/green] Excel saved to: [white]"{output_path}"[/white]')

    else:  # html
        from pystylometry.viz.jsx import export_bnc_frequency_jsx

        if args.input_dir:
            html_title = f"BNC Frequency Analysis: {args.input_dir.name} ({file_count} files)"
            html_source = str(args.input_dir)
        else:
            html_title = f"BNC Frequency Analysis: {args.input_file.name}"
            html_source = str(args.input_file)

        export_bnc_frequency_jsx(
            result,
            output_file=output_path,
            title=html_title,
            source_file=html_source,
        )

        abs_path = output_path.resolve()
        file_url = f"file://{abs_path}"
        console.print(f'[green]✓[/green] HTML report saved to: [white]"{output_path}"[/white]')
        console.print(f"  Open in browser: [link={file_url}]{file_url}[/link]")

    # Summary table
    console.print()
    table = Table(title="Summary", border_style="cyan", header_style="bold cyan")
    table.add_column("Metric", style="white")
    table.add_column("Count", justify="right", style="green")

    table.add_row("Total tokens", f"{result.total_tokens:,}")
    table.add_row("Unique words", f"{result.unique_tokens:,}")
    table.add_row("Not in BNC", f"[dim]{len(result.not_in_bnc):,}[/dim]")
    table.add_row("Underused", f"[blue]{len(result.underused):,}[/blue]")
    table.add_row("Overused", f"[red]{len(result.overused):,}[/red]")

    console.print(table)
    console.print()


def mega_cli() -> None:
    """CLI entry point for comprehensive mega analysis."""
    parser = argparse.ArgumentParser(
        prog="mega",
        description="Comprehensive stylometric analysis — every metric in one Excel workbook.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  mega --input-file manuscript.txt --output-file /tmp/analysis.xlsx
  mega --input-dir ~/ebooks/Author --output-file /tmp/author.xlsx

Generates a multi-tab Excel workbook with:
  Dashboard, Lexical Diversity, Readability, Sentences, Function Words,
  Style Markers, Character, N-grams, Prosody, Cohesion, Genre & Register,
  Dialect, Style Drift, Repetition, BNC Overused/Underused/Not Found.

Optional modules (readability, syntactic, prosody) are attempted and
gracefully skipped when dependencies are not installed.
""",
    )

    input_group = parser.add_mutually_exclusive_group(required=True)
    input_group.add_argument(
        "--input-file", "-i", type=Path, metavar="FILE",
        help="Path to text file to analyze",
    )
    input_group.add_argument(
        "--input-dir", "-d", type=Path, metavar="DIR",
        help="Directory of .txt files to analyze as a single corpus",
    )
    parser.add_argument(
        "--output-file", "-o", type=Path, required=True, metavar="FILE",
        help="Output Excel file path (.xlsx)",
    )

    args = parser.parse_args()

    from rich.console import Console
    from rich.panel import Panel
    from rich.text import Text

    console = Console(stderr=True)

    # ── Input loading ──
    if args.input_dir:
        if not args.input_dir.is_dir():
            console.print(f"[red]Error:[/red] Directory not found: {args.input_dir}")
            sys.exit(1)
        txt_files = sorted(args.input_dir.glob("*.txt"))
        if not txt_files:
            console.print(f"[red]Error:[/red] No .txt files found in: {args.input_dir}")
            sys.exit(1)
        parts: list[str] = []
        for tf in txt_files:
            try:
                parts.append(tf.read_text(encoding="utf-8"))
            except Exception as e:
                console.print(f"[red]Error reading {tf.name}:[/red] {e}")
                sys.exit(1)
        text = "\n".join(parts)
        file_count = len(txt_files)
    else:
        if not args.input_file.exists():
            console.print(f"[red]Error:[/red] File not found: {args.input_file}")
            sys.exit(1)
        try:
            text = args.input_file.read_text(encoding="utf-8")
        except Exception as e:
            console.print(f"[red]Error reading file:[/red] {e}")
            sys.exit(1)
        file_count = 1

    token_count = len(text.split())
    char_count = len(text)

    # ── Banner ──
    console.print()
    header = Text()
    header.append("PYSTYLOMETRY", style="bold cyan")
    header.append(" — ", style="dim")
    header.append("Mega Analysis", style="bold white")
    console.print(Panel(header, border_style="cyan"))

    console.print()
    console.print("[bold]INPUT[/bold]", style="cyan")
    console.print("─" * 60, style="dim")
    if args.input_dir:
        console.print(f"  Dir:     [white]{args.input_dir}[/white]")
        console.print(f"  Files:   [green]{file_count}[/green] .txt files")
    else:
        console.print(f"  File:    [white]{args.input_file}[/white]")
    console.print(f"  Size:    [green]{char_count:,}[/green] chars / [green]{token_count:,}[/green] tokens")

    console.print()
    console.print("[bold]OUTPUT[/bold]", style="cyan")
    console.print("─" * 60, style="dim")
    console.print(f"  Format:  [magenta]Excel (.xlsx)[/magenta]")
    console.print(f"  File:    [white]{args.output_file}[/white]")

    console.print()
    console.print("[bold]ANALYSIS[/bold]", style="cyan")
    console.print("─" * 60, style="dim")

    # ── Run all analyses ──
    from pystylometry.mega import run_mega, write_mega_excel

    results = run_mega(text, console)

    # ── Write Excel ──
    console.print()
    console.print("  [dim]Writing Excel workbook...[/dim]", end="")
    write_mega_excel(results, args.output_file)
    console.print(" [green]done[/green]")

    # ── Summary ──
    skipped = results.get("_skipped", [])
    tab_count = 17 - len(skipped)
    console.print()
    console.print(f'[green]✓[/green] Saved to: [white]"{args.output_file}"[/white]')
    console.print(f"  Tabs: [green]{tab_count}[/green] populated, [yellow]{len(skipped)}[/yellow] skipped")
    if skipped:
        for name, reason in skipped:
            console.print(f"    [yellow]•[/yellow] {name}: {reason}")
    console.print()


# ---------------------------------------------------------------------------
# word-class  —  Word classification distribution (Issue #54)
# ---------------------------------------------------------------------------


def classify_cli() -> None:
    """CLI entry point for word-class distribution analysis.

    Classifies every word in a text using the morphological taxonomy and
    produces a distribution table showing each classification label, its
    count, and its percentage of total words.

    Related GitHub Issue:
        #54 -- Add word-class distribution report
        https://github.com/craigtrim/pystylometry/issues/54
    """
    parser = argparse.ArgumentParser(
        prog="word-class",
        description="Produce a distribution table of word classifications.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  word-class --input-file manuscript.txt
  word-class -i manuscript.txt --format excel -o report.xlsx
  word-class --input-dir ~/ebooks/Author -o report.xlsx --format excel

Output:
  A table with one row per classification label (sorted a-z):
    label           count   percentage
    lexical         62134   96.8217

  Percentages should sum to 100%. If they don't, it indicates
  words not yet covered by the classification taxonomy.
""",
    )

    input_group = parser.add_mutually_exclusive_group(required=True)
    input_group.add_argument(
        "--input-file",
        "-i",
        type=Path,
        metavar="FILE",
        help="Path to text file to analyze",
    )
    input_group.add_argument(
        "--input-dir",
        "-d",
        type=Path,
        metavar="DIR",
        help="Directory of .txt files to analyze as a single corpus (requires --output-file)",
    )
    parser.add_argument(
        "--output-file",
        "-o",
        type=Path,
        default=None,
        metavar="FILE",
        help="Output file (default: <input>_word_class.<ext>)",
    )
    parser.add_argument(
        "--format",
        choices=["csv", "excel"],
        default="csv",
        help="Output format: csv (tab-delimited .tsv) or excel (.xlsx). Default: csv",
    )

    args = parser.parse_args()

    from rich.console import Console
    from rich.text import Text
    from rich.panel import Panel

    console = Console()

    # -- Input loading (matches BNC pattern) --------------------------------
    file_count: int = 1

    if args.input_dir:
        if not args.input_dir.is_dir():
            console.print(f"[red]Error:[/red] Directory not found: {args.input_dir}")
            sys.exit(1)
        if not args.output_file:
            console.print("[red]Error:[/red] --output-file is required when using --input-dir")
            sys.exit(1)

        txt_files = sorted(args.input_dir.glob("*.txt"))
        if not txt_files:
            console.print(f"[red]Error:[/red] No .txt files found in: {args.input_dir}")
            sys.exit(1)

        parts: list[str] = []
        for tf in txt_files:
            try:
                parts.append(tf.read_text(encoding="utf-8"))
            except Exception as e:
                console.print(f"[red]Error reading {tf.name}:[/red] {e}")
                sys.exit(1)
        text = "\n".join(parts)
        file_count = len(txt_files)
    else:
        if not args.input_file.exists():
            console.print(f"[red]Error:[/red] File not found: {args.input_file}")
            sys.exit(1)
        try:
            text = args.input_file.read_text(encoding="utf-8")
        except Exception as e:
            console.print(f"[red]Error reading file:[/red] {e}")
            sys.exit(1)

    # -- Determine output path ----------------------------------------------
    suffix_map = {"csv": ".tsv", "excel": ".xlsx"}
    if args.output_file:
        output_path = args.output_file
    else:
        suffix = suffix_map[args.format]
        output_path = args.input_file.with_name(
            f"{args.input_file.stem}_word_class{suffix}"
        )

    # -- Banner -------------------------------------------------------------
    token_count = len(text.split())
    char_count = len(text)

    console.print()
    header = Text()
    header.append("PYSTYLOMETRY", style="bold cyan")
    header.append(" — ", style="dim")
    header.append("Word Classification Distribution", style="bold white")
    console.print(Panel(header, border_style="cyan"))

    console.print()
    console.print("[bold]INPUT[/bold]", style="cyan")
    console.print("─" * 60, style="dim")
    if args.input_dir:
        console.print(f"  Dir:     [white]{args.input_dir}[/white]")
        console.print(f"  Files:   [green]{file_count}[/green] .txt files")
    else:
        console.print(f"  File:    [white]{args.input_file}[/white]")
    _sz = f"[green]{char_count:,}[/green] chars / [green]{token_count:,}[/green] tokens"
    console.print(f"  Size:    {_sz}")

    # -- Analyse ------------------------------------------------------------
    console.print()
    console.print("[bold]ANALYSIS[/bold]", style="cyan")
    console.print("─" * 60, style="dim")

    from pystylometry.lexical.word_class_distribution import (
        compute_word_class_distribution,
    )

    with console.status("  Classifying words..."):
        result = compute_word_class_distribution(text)

    console.print(f"  Tokens:  [green]{result.total_words:,}[/green]")
    console.print(f"  Labels:  [green]{result.unique_labels}[/green] distinct classifications")

    pct_sum = result.percentage_sum
    if abs(pct_sum - 100.0) > 0.01:
        console.print(
            f"  Sum:     [yellow]{pct_sum:.4f}%[/yellow]  (expected 100% — "
            f"gap indicates unclassified words)"
        )
    else:
        console.print(f"  Sum:     [green]{pct_sum:.4f}%[/green]")

    # -- Write output -------------------------------------------------------
    console.print()
    console.print("[bold]OUTPUT[/bold]", style="cyan")
    console.print("─" * 60, style="dim")

    if args.format == "csv":
        lines = [
            "label\tcount\tpercentage\truns\trun_mean\trun_median"
            "\trun_mode\trun_min\trun_max"
        ]
        for entry in result.classifications:
            rs = entry.run_stats
            lines.append(
                f"{entry.label}\t{entry.count}\t{entry.percentage:.4f}"
                f"\t{rs.runs}\t{rs.mean:.4f}\t{rs.median:.4f}"
                f"\t{rs.mode}\t{rs.min}\t{rs.max}"
            )
        output_path.write_text("\n".join(lines) + "\n", encoding="utf-8")
        console.print(f"  Format:  [magenta]CSV (tab-delimited .tsv)[/magenta]")

    elif args.format == "excel":
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font, PatternFill
        except ImportError:
            console.print("[red]Error:[/red] Excel export requires openpyxl.")
            console.print("  Install with: [yellow]pip install pystylometry[excel][/yellow]")
            console.print("  Or for pipx: [yellow]pipx inject pystylometry openpyxl[/yellow]")
            sys.exit(1)

        wb = Workbook()
        wb.remove(wb.active)

        align = Alignment(horizontal="center", vertical="center")
        font_header = Font(bold=True, size=14)
        fill_header = PatternFill(
            start_color="D9E2F3", end_color="D9E2F3", fill_type="solid"
        )
        font_bold = Font(bold=True)

        ws = wb.create_sheet("Word Classification")
        ws.append([
            "Label", "Count", "Unique", "Percentage",
            "Runs", "Run Mean", "Run Median", "Run Mode", "Run Min", "Run Max",
        ])

        # Style header
        for cell in ws[1]:
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = align

        # Data rows
        for entry in result.classifications:
            rs = entry.run_stats
            ws.append([
                entry.label, entry.count, entry.unique,
                round(entry.percentage, 4),
                rs.runs, round(rs.mean, 4), round(rs.median, 4),
                rs.mode, rs.min, rs.max,
            ])

        # Summary row
        total_unique = sum(e.unique for e in result.classifications)
        ws.append([])
        ws.append(["TOTAL", result.total_words, total_unique, round(pct_sum, 4)])
        summary_row = ws.max_row
        for col in range(1, 11):
            ws.cell(row=summary_row, column=col).font = font_bold
            ws.cell(row=summary_row, column=col).alignment = align

        # Warning if sum != 100%
        if abs(pct_sum - 100.0) > 0.01:
            ws.append([])
            ws.append([f"WARNING: Sum is {pct_sum:.4f}%, expected 100%"])
            ws.cell(row=ws.max_row, column=1).font = Font(color="FF0000", bold=True)

        # Column widths and alignment
        ws.column_dimensions["A"].width = 40
        for letter in ("B", "C", "D", "E", "F", "G", "H", "I", "J"):
            ws.column_dimensions[letter].width = 12
        # Integer columns get comma formatting: Count(B), Unique(C), Runs(E),
        # Run Mode(H), Run Min(I), Run Max(J).
        comma_cols = {2, 3, 5, 8, 9, 10}
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.alignment = align
                if cell.column in comma_cols and isinstance(cell.value, (int, float)):
                    cell.number_format = "#,##0"
        # Left-align the label column
        for row in ws.iter_rows(min_col=1, max_col=1):
            for cell in row:
                if cell.row > 1:
                    cell.alignment = Alignment(horizontal="left", vertical="center")

        # Row heights
        ws.row_dimensions[1].height = 25
        for r in range(2, ws.max_row + 1):
            ws.row_dimensions[r].height = 20

        # ── Max Run Examples tab (labels where run_max > 1 and < 10) ──
        if result.max_run_examples:
            from collections import Counter as _ExCounter

            ws_ex = wb.create_sheet("Run Max Examples")
            ws_ex.append(["Label", "Run Max", "Tokens", "Occurrences"])
            for cell in ws_ex[1]:
                cell.font = font_header
                cell.fill = fill_header
                cell.alignment = align

            for label in sorted(result.max_run_examples):
                examples = result.max_run_examples[label]
                entry = next(e for e in result.classifications if e.label == label)
                # Deduplicate: count how many times each unique sequence appears
                seq_counts = _ExCounter(tuple(seq) for seq in examples)
                for seq_tuple, occ in sorted(
                    seq_counts.items(), key=lambda x: (-x[1], x[0]),
                ):
                    ws_ex.append([
                        label, entry.run_stats.max,
                        " ".join(seq_tuple), occ,
                    ])

            ws_ex.column_dimensions["A"].width = 40
            ws_ex.column_dimensions["B"].width = 12
            ws_ex.column_dimensions["C"].width = 60
            ws_ex.column_dimensions["D"].width = 14
            for row in ws_ex.iter_rows(min_row=2, max_row=ws_ex.max_row):
                for cell in row:
                    cell.alignment = align
            # Left-align label and tokens columns
            for row in ws_ex.iter_rows(min_col=1, max_col=1, min_row=2):
                for cell in row:
                    cell.alignment = Alignment(horizontal="left", vertical="center")
            for row in ws_ex.iter_rows(min_col=3, max_col=3, min_row=2):
                for cell in row:
                    cell.alignment = Alignment(horizontal="left", vertical="center")
            ws_ex.row_dimensions[1].height = 25
            for r in range(2, ws_ex.max_row + 1):
                ws_ex.row_dimensions[r].height = 20

        # ── Tokens by Label tab (unique tokens + count, excludes lexical) ──
        if result.tokens_by_label:
            from collections import Counter as _Counter
            from openpyxl.formatting.rule import ColorScaleRule
            from openpyxl.styles import Border, Side
            from openpyxl.utils import get_column_letter

            import re as _re

            ws_tok = wb.create_sheet("Tokens by Label")
            sorted_labels = sorted(result.tokens_by_label)

            def _break_label(lbl: str) -> str:
                return "\n".join(_re.split(r"[._]", lbl))

            # ── Header row ──
            header: list[str] = []
            for label in sorted_labels:
                header.append(_break_label(label))
                header.append("N")
            ws_tok.append(header)

            # Fonts
            font_label_hdr = Font(name="Calibri", size=14, italic=True)
            font_n_hdr = Font(name="Calibri", size=12, italic=True)
            font_word = Font(name="Calibri", size=11)
            font_num = Font(name="Calibri", size=10)

            # Alignments
            hdr_align_token = Alignment(
                horizontal="right", vertical="bottom", wrap_text=True,
            )
            hdr_align_n = Alignment(
                horizontal="left", vertical="bottom", wrap_text=True,
            )
            align_token = Alignment(horizontal="right", vertical="center")
            align_count = Alignment(horizontal="left", vertical="center")

            # Fills
            fill_white = PatternFill(
                start_color="FFFFFF", end_color="FFFFFF", fill_type="solid",
            )

            # Border side
            thin = Side(style="thin")

            # Style headers
            for cell in ws_tok[1]:
                cell.fill = fill_header
                if cell.column % 2 == 1:
                    cell.font = font_label_hdr
                    cell.alignment = hdr_align_token
                else:
                    cell.font = font_n_hdr
                    cell.alignment = hdr_align_n

            # Header row height
            max_segments = max(
                len(_re.split(r"[._]", lbl)) for lbl in sorted_labels
            )
            ws_tok.row_dimensions[1].height = max_segments * 15 + 10

            # ── Deduplicate tokens per label ──
            label_unique: dict[str, list[tuple[str, int]]] = {}
            max_unique = 0
            for label in sorted_labels:
                counts = _Counter(result.tokens_by_label[label])
                ranked = sorted(counts.items(), key=lambda x: (-x[1], x[0]))
                label_unique[label] = ranked
                max_unique = max(max_unique, len(ranked))

            # ── Column widths ──
            total_cols = len(sorted_labels) * 2
            for col_idx in range(1, total_cols + 1):
                letter = get_column_letter(col_idx)
                ws_tok.column_dimensions[letter].width = (
                    25 if col_idx % 2 == 1 else 10
                )

            # ── Write data + style per column pair ──
            for pair_idx, label in enumerate(sorted_labels):
                col_token = pair_idx * 2 + 1
                col_count = pair_idx * 2 + 2
                items = label_unique[label]
                n_items = len(items)
                last_data_row = n_items + 1  # 1-based (header is row 1)

                # Write values
                for row_idx, (token, cnt) in enumerate(items):
                    excel_row = row_idx + 2
                    ws_tok.cell(
                        row=excel_row, column=col_token, value=token,
                    )
                    ws_tok.cell(
                        row=excel_row, column=col_count, value=cnt,
                    )

                # Style data cells + outline borders per column
                for row_idx in range(n_items):
                    excel_row = row_idx + 2
                    is_first = row_idx == 0
                    is_last = row_idx == n_items - 1

                    # Determine border edges for this row in the column
                    top_side = thin if is_first else None
                    bottom_side = thin if is_last else None

                    bdr = Border(
                        left=thin, right=thin,
                        top=top_side, bottom=bottom_side,
                    )

                    cell_tok = ws_tok.cell(row=excel_row, column=col_token)
                    cell_tok.font = font_word
                    cell_tok.alignment = align_token
                    cell_tok.fill = fill_white
                    cell_tok.border = bdr

                    cell_cnt = ws_tok.cell(row=excel_row, column=col_count)
                    cell_cnt.font = font_num
                    cell_cnt.alignment = align_count
                    cell_cnt.fill = fill_white
                    cell_cnt.border = bdr

                # Header outline borders (top + left/right + bottom)
                hdr_bdr_tok = Border(
                    left=thin, right=thin, top=thin, bottom=thin,
                )
                ws_tok.cell(row=1, column=col_token).border = hdr_bdr_tok
                ws_tok.cell(row=1, column=col_count).border = hdr_bdr_tok

                # Conditional formatting: green→white on count column
                if n_items > 0:
                    cnt_letter = get_column_letter(col_count)
                    cell_range = f"{cnt_letter}2:{cnt_letter}{last_data_row}"
                    ws_tok.conditional_formatting.add(
                        cell_range,
                        ColorScaleRule(
                            start_type="min", start_color="FFFFFF",
                            end_type="max", end_color="63BE7B",
                        ),
                    )

            # ── Row heights ──
            for row_idx in range(max_unique):
                ws_tok.row_dimensions[row_idx + 2].height = 20

        wb.save(output_path)
        console.print(f"  Format:  [magenta]Excel (.xlsx)[/magenta]")

    console.print(f"  File:    [white]{output_path}[/white]")
    console.print()
    console.print(f'[green]✓[/green] Saved to: [white]"{output_path}"[/white]')
    console.print()


# ---------------------------------------------------------------------------
# bnc-meta  —  Comparative BNC summary analysis across authors (Issue #55)
# ---------------------------------------------------------------------------


def bnc_meta_cli() -> None:
    """CLI entry point for comparative BNC meta-analysis across authors.

    Reads summary statistics from multiple per-author BNC Excel files and
    produces a single comparative workbook with counts and percentages
    side-by-side.

    Related GitHub Issue:
        #55 -- Add bnc-meta CLI for comparative BNC summary analysis across authors
        https://github.com/craigtrim/pystylometry/issues/55
    """
    parser = argparse.ArgumentParser(
        prog="bnc-meta",
        description="Compare BNC summary statistics across multiple author Excel files.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  bnc-meta --input-dir /tmp --output-file /tmp/bnc-comparison.xlsx
  bnc-meta -d /tmp -o /tmp/comparison.xlsx

Input:
  A directory containing .xlsx files produced by:
    bnc --input-dir [author-dir] --format excel --output-file Author.xlsx

  Each file must have a "summary" sheet with BNC stats.
  Author name is derived from the filename (e.g., "Mark Twain.xlsx" -> "Mark Twain").

Output:
  An Excel workbook with two sheets:
    counts       - Raw counts (Total Tokens, Unique Tokens, Overused, Underused, Not in BNC)
    percentages  - Percentage of total tokens (Unique, Overused, Underused, Not in BNC)
""",
    )

    parser.add_argument(
        "--input-dir",
        "-d",
        type=Path,
        required=True,
        metavar="DIR",
        help="Directory of BNC Excel files (one per author)",
    )
    parser.add_argument(
        "--output-file",
        "-o",
        type=Path,
        required=True,
        metavar="FILE",
        help="Output Excel file path (.xlsx)",
    )

    args = parser.parse_args()

    from rich.console import Console
    from rich.panel import Panel
    from rich.table import Table
    from rich.text import Text

    console = Console(stderr=True)

    # ── Validation ──
    if not args.input_dir.is_dir():
        console.print(f"[red]Error:[/red] Directory not found: {args.input_dir}")
        sys.exit(1)

    xlsx_files = sorted(args.input_dir.glob("*.xlsx"))
    xlsx_files = [f for f in xlsx_files if not f.name.startswith("~$")]
    if not xlsx_files:
        console.print(f"[red]Error:[/red] No .xlsx files found in: {args.input_dir}")
        sys.exit(1)

    # ── Banner ──
    console.print()
    header = Text()
    header.append("PYSTYLOMETRY", style="bold cyan")
    header.append(" — ", style="dim")
    header.append("BNC Meta-Analysis", style="bold white")
    console.print(Panel(header, border_style="cyan"))

    console.print()
    console.print("[bold]INPUT[/bold]", style="cyan")
    console.print("─" * 60, style="dim")
    console.print(f"  Dir:     [white]{args.input_dir}[/white]")
    console.print(f"  Files:   [green]{len(xlsx_files)}[/green] .xlsx files")

    console.print()
    console.print("[bold]OUTPUT[/bold]", style="cyan")
    console.print("─" * 60, style="dim")
    console.print(f"  Format:  [magenta]Excel (.xlsx)[/magenta]")
    console.print(f"  File:    [white]{args.output_file}[/white]")

    # ── Read summaries ──
    console.print()
    console.print("[bold]ANALYSIS[/bold]", style="cyan")
    console.print("─" * 60, style="dim")

    from pystylometry.lexical.bnc_meta import (
        METRIC_LABELS,
        read_bnc_summaries,
        write_bnc_meta_excel,
    )

    with console.status("  Reading BNC Excel files..."):
        summaries = read_bnc_summaries(args.input_dir)

    if not summaries:
        console.print("[red]Error:[/red] No valid BNC summary sheets found in any .xlsx file.")
        sys.exit(1)

    skipped = len(xlsx_files) - len(summaries)
    console.print(f"  Authors: [green]{len(summaries)}[/green] loaded")
    if skipped:
        console.print(f"  Skipped: [yellow]{skipped}[/yellow] (missing summary sheet)")

    # ── Write Excel ──
    with console.status("  Writing comparative workbook..."):
        write_bnc_meta_excel(summaries, args.output_file)

    # ── Console preview (authors as rows, truncated for large sets) ──
    console.print()
    max_preview = 10
    preview = summaries[:max_preview]

    table = Table(title="Counts", border_style="cyan", header_style="bold cyan")
    table.add_column("Author", style="bold")
    for metric in METRIC_LABELS:
        table.add_column(metric, justify="right")

    for s in preview:
        table.add_row(
            s.author, f"{s.total_tokens:,}", f"{s.unique_tokens:,}",
            f"{s.overused:,}", f"{s.underused:,}", f"{s.not_in_bnc:,}",
        )
    if len(summaries) > max_preview:
        table.add_row(f"... +{len(summaries) - max_preview} more", *[""] * len(METRIC_LABELS), style="dim")

    console.print(table)

    console.print()
    console.print(f'[green]✓[/green] Saved to: [white]"{args.output_file}"[/white]')
    console.print(f"  Sheets: [green]counts[/green], [green]percentages[/green]")
    console.print(f"  Authors: [green]{len(summaries)}[/green] columns")
    console.print()


def mega_meta_cli() -> None:
    """CLI entry point for comparative mega meta-analysis across authors.

    Reads N-gram and Prosody metrics from multiple per-author mega Excel files
    and produces a single comparative workbook.
    """
    parser = argparse.ArgumentParser(
        prog="mega-meta",
        description="Compare mega stylometric metrics across multiple author Excel files.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  mega-meta --input-dir ~/Desktop/mega-output --output-file /tmp/mega-comparison.xlsx
  mega-meta -d ~/Desktop/mega-output -o /tmp/mega-comparison.xlsx

Input:
  A directory (with optional subdirectories) containing .xlsx files produced by:
    mega --input-dir [author-dir] --output-file Author.xlsx

  Each file should have an "N-grams" and/or "Prosody" sheet.
  Author name is derived from the filename (e.g., "Mark Twain.xlsx" -> "Mark Twain").

Output:
  An Excel workbook with N-grams and Prosody sheets (authors as rows, metrics as columns).
""",
    )

    parser.add_argument(
        "--input-dir",
        "-d",
        type=Path,
        required=True,
        metavar="DIR",
        help="Directory of mega Excel files (one per author)",
    )
    parser.add_argument(
        "--output-file",
        "-o",
        type=Path,
        required=True,
        metavar="FILE",
        help="Output Excel file path (.xlsx)",
    )

    args = parser.parse_args()

    from rich.console import Console
    from rich.panel import Panel
    from rich.table import Table
    from rich.text import Text

    console = Console(stderr=True)

    # ── Validation ──
    if not args.input_dir.is_dir():
        console.print(f"[red]Error:[/red] Directory not found: {args.input_dir}")
        sys.exit(1)

    # ── Banner ──
    console.print()
    header = Text()
    header.append("PYSTYLOMETRY", style="bold cyan")
    header.append(" — ", style="dim")
    header.append("Mega Meta-Analysis", style="bold white")
    console.print(Panel(header, border_style="cyan"))

    # ── Read summaries ──
    console.print()
    console.print("[bold]INPUT[/bold]", style="cyan")
    console.print("─" * 60, style="dim")
    console.print(f"  Dir:     [white]{args.input_dir}[/white]")

    from pystylometry.lexical.mega_meta import (
        CHARACTER_METRIC_LABELS,
        NGRAM_METRIC_LABELS,
        PROSODY_METRIC_LABELS,
        SYLLABLE_BUCKETS,
        read_character_summaries,
        read_ngram_summaries,
        read_prosody_summaries,
        read_syllable_summaries,
        write_mega_meta_excel,
    )

    with console.status("  Reading mega Excel files..."):
        summaries = read_ngram_summaries(args.input_dir)
        prosody_summaries = read_prosody_summaries(args.input_dir)
        syllable_summaries = read_syllable_summaries(args.input_dir)
        character_summaries = read_character_summaries(args.input_dir)

    if not summaries and not prosody_summaries and not syllable_summaries and not character_summaries:
        console.print("[red]Error:[/red] No valid sheets found in any .xlsx file.")
        sys.exit(1)

    console.print(f"  N-grams:    [green]{len(summaries)}[/green] authors loaded")
    console.print(f"  Prosody:    [green]{len(prosody_summaries)}[/green] authors loaded")
    console.print(f"  Syllables:  [green]{len(syllable_summaries)}[/green] authors loaded")
    console.print(f"  Character:  [green]{len(character_summaries)}[/green] authors loaded")

    console.print()
    console.print("[bold]OUTPUT[/bold]", style="cyan")
    console.print("─" * 60, style="dim")
    console.print(f"  Format:  [magenta]Excel (.xlsx)[/magenta]")
    console.print(f"  File:    [white]{args.output_file}[/white]")

    # ── Write Excel ──
    with console.status("  Writing comparative workbook..."):
        write_mega_meta_excel(
            summaries, args.output_file,
            prosody_summaries=prosody_summaries,
            syllable_summaries=syllable_summaries,
            character_summaries=character_summaries,
        )

    # ── Console preview: N-grams ──
    console.print()
    max_preview = 10

    if summaries:
        preview = summaries[:max_preview]
        table = Table(title="N-grams", border_style="cyan", header_style="bold cyan")
        table.add_column("Author", style="bold")
        for metric in NGRAM_METRIC_LABELS:
            table.add_column(metric, justify="right")

        for s in preview:
            table.add_row(
                s.author,
                f"{s.char_bigram_entropy:.4f}",
                f"{s.char_bigram_perplexity:.4f}",
                f"{s.word_bigram_entropy:.4f}",
                f"{s.word_bigram_perplexity:,}",
            )
        if len(summaries) > max_preview:
            table.add_row(
                f"... +{len(summaries) - max_preview} more",
                *[""] * len(NGRAM_METRIC_LABELS),
                style="dim",
            )
        console.print(table)

    # ── Console preview: Prosody ──
    if prosody_summaries:
        preview_p = prosody_summaries[:max_preview]
        table_p = Table(title="Prosody", border_style="cyan", header_style="bold cyan")
        table_p.add_column("Author", style="bold")
        for metric in PROSODY_METRIC_LABELS:
            table_p.add_column(metric, justify="right")

        for p in preview_p:
            table_p.add_row(
                p.author,
                f"{p.mean_syllables_per_word:.4f}",
                f"{p.syllable_std_dev:.4f}",
                f"{p.polysyllabic_ratio:.4f}",
                f"{p.monosyllabic_ratio:.4f}",
                f"{p.rhythmic_regularity:.4f}",
                f"{p.alliteration_density:.4f}",
                f"{p.assonance_density:.4f}",
                f"{p.consonance_density:.4f}",
                f"{p.sentence_rhythm_score:.4f}",
                f"{p.avg_consonant_cluster:.4f}",
            )
        if len(prosody_summaries) > max_preview:
            table_p.add_row(
                f"... +{len(prosody_summaries) - max_preview} more",
                *[""] * len(PROSODY_METRIC_LABELS),
                style="dim",
            )
        console.print()
        console.print(table_p)

    # ── Console preview: Syllables ──
    if syllable_summaries:
        preview_s = syllable_summaries[:max_preview]
        table_s = Table(title="Syllables", border_style="cyan", header_style="bold cyan")
        table_s.add_column("Author", style="bold")
        for bucket in SYLLABLE_BUCKETS:
            table_s.add_column(bucket, justify="right")

        for s in preview_s:
            table_s.add_row(
                s.author,
                *[f"{s.totals.get(b, 0):,}" for b in SYLLABLE_BUCKETS],
            )
        if len(syllable_summaries) > max_preview:
            table_s.add_row(
                f"... +{len(syllable_summaries) - max_preview} more",
                *[""] * len(SYLLABLE_BUCKETS),
                style="dim",
            )
        console.print()
        console.print(table_s)

    # ── Console preview: Character ──
    if character_summaries:
        preview_c = character_summaries[:max_preview]
        table_c = Table(title="Character", border_style="cyan", header_style="bold cyan")
        table_c.add_column("Author", style="bold")
        for metric in CHARACTER_METRIC_LABELS:
            table_c.add_column(metric, justify="right")

        for c in preview_c:
            table_c.add_row(
                c.author,
                f"{c.avg_word_length:.4f}",
                f"{c.avg_sentence_length_chars:.4f}",
                f"{c.punctuation_density:.4f}",
                f"{c.punctuation_variety:.4f}",
                f"{c.vowel_consonant_ratio:.4f}",
                f"{c.digit_ratio:.4f}",
                f"{c.uppercase_ratio:.4f}",
                f"{c.whitespace_ratio:.4f}",
            )
        if len(character_summaries) > max_preview:
            table_c.add_row(
                f"... +{len(character_summaries) - max_preview} more",
                *[""] * len(CHARACTER_METRIC_LABELS),
                style="dim",
            )
        console.print()
        console.print(table_c)

    # ── Summary ──
    console.print()
    console.print(f'[green]✓[/green] Saved to: [white]"{args.output_file}"[/white]')
    sheets = []
    if summaries:
        sheets.append(f"N-grams ({len(summaries)} authors)")
    if prosody_summaries:
        sheets.append(f"Prosody ({len(prosody_summaries)} authors)")
    if syllable_summaries:
        sheets.append(f"Syllables ({len(syllable_summaries)} authors)")
    if character_summaries:
        sheets.append(f"Character ({len(character_summaries)} authors)")
    console.print(f"  Sheets: [green]{', '.join(sheets)}[/green]")
    console.print()


if __name__ == "__main__":
    drift_cli()
