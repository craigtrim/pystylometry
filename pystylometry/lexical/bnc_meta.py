"""BNC meta-analysis: comparative summary across multiple author BNC Excel files.

Reads the summary sheet from each per-author BNC Excel file (as produced by
``bnc --format excel``) and aggregates counts and percentages into
side-by-side comparative tables.

Usage (via CLI):
    bnc-meta --input-dir /tmp --output-file /tmp/bnc-comparison.xlsx

Related GitHub Issue:
    #55 -- Add bnc-meta CLI for comparative BNC summary analysis across authors
    https://github.com/craigtrim/pystylometry/issues/55
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------
METRIC_LABELS = ["Total Tokens", "Unique Tokens", "Overused", "Underused", "Not in BNC"]
PCT_METRIC_LABELS = ["Unique Tokens", "Overused", "Underused", "Not in BNC"]

_ALIGN = Alignment(horizontal="center", vertical="center")
_ALIGN_LEFT = Alignment(horizontal="left", vertical="center")
_FONT_HEADER = Font(bold=True, size=14)
_FILL_HEADER = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")


# ---------------------------------------------------------------------------
# Data
# ---------------------------------------------------------------------------
@dataclass
class AuthorSummary:
    """Summary stats for one author, extracted from a BNC Excel file."""

    author: str
    total_tokens: int
    unique_tokens: int
    overused: int
    underused: int
    not_in_bnc: int
    pct_unique_tokens: float
    pct_overused: float
    pct_underused: float
    pct_not_in_bnc: float


def _parse_int(s: str) -> int:
    """Parse a formatted integer string like '3,709,051' to int."""
    try:
        return int(s.replace(",", ""))
    except (ValueError, AttributeError):
        return 0


def _parse_pct(s: str) -> float:
    """Parse a percentage string like '1.86%' to float (e.g. 1.86)."""
    try:
        return float(s.replace("%", "").strip())
    except (ValueError, AttributeError):
        return 0.0


# ---------------------------------------------------------------------------
# Read
# ---------------------------------------------------------------------------
def _find_stats_header_row(ws: object) -> int | None:
    """Scan for the row containing the 'Metric' header text."""
    for row in ws.iter_rows(min_row=1, max_row=30, min_col=1, max_col=1):
        for cell in row:
            if cell.value and str(cell.value).strip() == "Metric":
                return cell.row
    return None


def read_author_summary(xlsx_path: Path) -> AuthorSummary | None:
    """Read summary stats from a single BNC Excel file.

    Returns ``None`` if the file cannot be parsed (missing summary sheet,
    unexpected layout, etc.).
    """
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        if "summary" not in wb.sheetnames:
            return None
        ws = wb["summary"]

        header_row = _find_stats_header_row(ws)
        if header_row is None:
            return None

        # Stats occupy the 5 rows immediately after the header row
        def _cell(row_offset: int, col: int) -> str:
            val = ws.cell(row=header_row + row_offset, column=col).value
            return str(val).strip() if val is not None else ""

        return AuthorSummary(
            author=xlsx_path.stem,
            total_tokens=_parse_int(_cell(1, 2)),
            unique_tokens=_parse_int(_cell(2, 2)),
            overused=_parse_int(_cell(3, 2)),
            underused=_parse_int(_cell(4, 2)),
            not_in_bnc=_parse_int(_cell(5, 2)),
            pct_unique_tokens=_parse_pct(_cell(2, 3)),
            pct_overused=_parse_pct(_cell(3, 3)),
            pct_underused=_parse_pct(_cell(4, 3)),
            pct_not_in_bnc=_parse_pct(_cell(5, 3)),
        )
    finally:
        wb.close()


def read_bnc_summaries(input_dir: Path) -> list[AuthorSummary]:
    """Read all BNC Excel files from a directory and return sorted summaries."""
    summaries: list[AuthorSummary] = []
    for xlsx_path in sorted(input_dir.glob("*.xlsx")):
        # Skip temporary/lock files
        if xlsx_path.name.startswith("~$"):
            continue
        result = read_author_summary(xlsx_path)
        if result is not None:
            summaries.append(result)
    return sorted(summaries, key=lambda s: s.author.lower())


# ---------------------------------------------------------------------------
# Write
# ---------------------------------------------------------------------------
def _auto_width(ws: object, min_w: int = 15, max_w: int = 40) -> None:
    """Set reasonable column widths."""
    for col in ws.columns:
        letter = col[0].column_letter
        lengths = [len(str(c.value)) for c in col if c.value is not None]
        best = max(lengths) if lengths else min_w
        ws.column_dimensions[letter].width = min(max(best + 2, min_w), max_w)


def write_bnc_meta_excel(summaries: list[AuthorSummary], output_path: Path) -> None:
    """Write comparative BNC meta-analysis to an Excel workbook.

    Creates two sheets:
      - ``counts``: raw token counts per author
      - ``percentages``: percentage of total tokens per author
    """
    wb = Workbook()
    wb.remove(wb.active)

    authors = [s.author for s in summaries]

    # -- Counts sheet (authors as rows, metrics as columns) -----------------
    ws_counts = wb.create_sheet("counts")
    ws_counts.append(["Author"] + METRIC_LABELS)

    for s in summaries:
        ws_counts.append([
            s.author, s.total_tokens, s.unique_tokens,
            s.overused, s.underused, s.not_in_bnc,
        ])

    # -- Percentages sheet (authors as rows, metrics as columns) -----------
    ws_pct = wb.create_sheet("percentages")
    ws_pct.append(["Author"] + PCT_METRIC_LABELS)

    for s in summaries:
        ws_pct.append([
            s.author, s.pct_unique_tokens, s.pct_overused,
            s.pct_underused, s.pct_not_in_bnc,
        ])

    # -- Styling ------------------------------------------------------------
    for ws in [ws_counts, ws_pct]:
        # Header row
        for cell in ws[1]:
            cell.font = _FONT_HEADER
            cell.fill = _FILL_HEADER
            cell.alignment = _ALIGN

        # Data cells
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = _ALIGN
            # Left-align the author name column
            row[0].alignment = _ALIGN_LEFT

        # Row heights
        ws.row_dimensions[1].height = 25
        for r in range(2, ws.max_row + 1):
            ws.row_dimensions[r].height = 20

        _auto_width(ws)

    # Number formats: counts get comma-separated integers
    for col_idx in range(2, 2 + len(METRIC_LABELS)):
        col_letter = get_column_letter(col_idx)
        for row_idx in range(2, ws_counts.max_row + 1):
            ws_counts.cell(row=row_idx, column=col_idx).number_format = "#,##0"

    # Percentages get 2-decimal percent format
    for col_idx in range(2, 2 + len(PCT_METRIC_LABELS)):
        for row_idx in range(2, ws_pct.max_row + 1):
            ws_pct.cell(row=row_idx, column=col_idx).number_format = "0.00\\%"

    wb.save(output_path)
