"""Mega meta-analysis: comparative summary across multiple mega Excel files.

Reads the N-grams and Prosody sheets from each per-author mega Excel file
(as produced by ``mega``) and aggregates metrics into side-by-side
comparative tables.

Usage (via CLI):
    mega-meta --input-dir ~/Desktop/mega-output --output-file /tmp/mega-comparison.xlsx
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------
NGRAM_METRIC_LABELS = [
    "Char Bigram Entropy",
    "Char Bigram Perplexity",
    "Word Bigram Entropy",
    "Word Bigram Perplexity",
]

PROSODY_METRIC_LABELS = [
    "Mean Syllables/Word",
    "Syllable Std Dev",
    "Polysyllabic Ratio (3+)",
    "Monosyllabic Ratio",
    "Rhythmic Regularity",
    "Alliteration Density",
    "Assonance Density",
    "Consonance Density",
    "Sentence Rhythm Score",
    "Avg Consonant Cluster",
]

# Syllable distribution buckets written by mega's Prosody sheet (1–11, 12+)
SYLLABLE_BUCKETS = [str(b) for b in range(1, 12)] + ["12+"]

CHARACTER_METRIC_LABELS = [
    "Avg Word Length",
    "Avg Sentence Length (chars)",
    "Punctuation Density",
    "Punctuation Variety",
    "Vowel:Consonant Ratio",
    "Digit Ratio",
    "Uppercase Ratio",
    "Whitespace Ratio",
]

LEXICAL_METRIC_LABELS = [
    "TTR",
    "Root TTR",
    "Log TTR",
    "STTR",
    "MTLD (forward)",
    "MTLD (backward)",
    "MTLD (average)",
    "Yule's K",
    "Yule's I",
    "Hapax Ratio",
    "Dis-Hapax Ratio",
    "Sichel's S",
    "Honore's R",
    "VocD-D",
    "MATTR",
    "HDD",
    "MSTTR",
]

BNC_METRIC_LABELS = ["Total Tokens", "Unique Tokens", "Overused", "Underused", "Not in BNC"]
BNC_PCT_METRIC_LABELS = ["Unique Tokens", "Overused", "Underused", "Not in BNC"]

_ALIGN = Alignment(horizontal="center", vertical="center")
_ALIGN_LEFT = Alignment(horizontal="left", vertical="center")
_FONT_HEADER = Font(bold=True, size=14)
_FILL_HEADER = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")


# ---------------------------------------------------------------------------
# Data
# ---------------------------------------------------------------------------
@dataclass
class NgramSummary:
    """N-gram entropy/perplexity stats for one author."""

    author: str
    char_bigram_entropy: float
    char_bigram_perplexity: float
    word_bigram_entropy: float
    word_bigram_perplexity: int


@dataclass
class ProsodySummary:
    """Prosody metrics for one author."""

    author: str
    mean_syllables_per_word: float
    syllable_std_dev: float
    polysyllabic_ratio: float
    monosyllabic_ratio: float
    rhythmic_regularity: float
    alliteration_density: float
    assonance_density: float
    consonance_density: float
    sentence_rhythm_score: float
    avg_consonant_cluster: float


@dataclass
class SyllableSummary:
    """Total syllable counts per bucket for one author.

    Buckets are "1" through "11" and "12+", matching the syllable
    distribution section of the mega Prosody sheet.
    """

    author: str
    totals: dict[str, int]  # bucket -> total count


@dataclass
class CharacterSummary:
    """Character-level metrics for one author."""

    author: str
    avg_word_length: float
    avg_sentence_length_chars: float
    punctuation_density: float
    punctuation_variety: float
    vowel_consonant_ratio: float
    digit_ratio: float
    uppercase_ratio: float
    whitespace_ratio: float


@dataclass
class LexicalDiversitySummary:
    """Lexical diversity metrics for one author."""

    author: str
    ttr: float | None = None
    root_ttr: float | None = None
    log_ttr: float | None = None
    sttr: float | None = None
    mtld_forward: float | None = None
    mtld_backward: float | None = None
    mtld_average: float | None = None
    yule_k: float | None = None
    yule_i: float | None = None
    hapax_ratio: float | None = None
    dis_hapax_ratio: float | None = None
    sichel_s: float | None = None
    honore_r: float | None = None
    vocd_d: float | None = None
    mattr: float | None = None
    hdd: float | None = None
    msttr: float | None = None


@dataclass
class BncSummary:
    """BNC frequency metrics for one author."""

    author: str
    total_tokens: int
    unique_tokens: int
    overused: int
    underused: int
    not_in_bnc: int
    pct_unique_tokens: float
    pct_overused: float
    pct_underused: float
    pct_not_in_bnc: float


# ---------------------------------------------------------------------------
# Read
# ---------------------------------------------------------------------------
_EXPECTED_METRICS = {
    "Char Bigram Entropy": "char_bigram_entropy",
    "Char Bigram Perplexity": "char_bigram_perplexity",
    "Word Bigram Entropy": "word_bigram_entropy",
    "Word Bigram Perplexity": "word_bigram_perplexity",
}


def read_ngram_summary(xlsx_path: Path) -> NgramSummary | None:
    """Read N-gram stats from a single mega Excel file.

    Returns ``None`` if the file cannot be parsed (missing N-grams sheet,
    unexpected layout, etc.).
    """
    try:
        wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    except Exception:
        # Skip invalid/corrupted Excel files
        return None
    try:
        if "N-grams" not in wb.sheetnames:
            return None
        ws = wb["N-grams"]

        values: dict[str, float] = {}
        for row in ws.iter_rows(min_row=1, max_row=10, min_col=1, max_col=2):
            label = row[0].value
            val = row[1].value if len(row) > 1 else None
            if label and str(label).strip() in _EXPECTED_METRICS and val is not None:
                try:
                    field = _EXPECTED_METRICS[str(label).strip()]
                    v = float(val)
                    values[field] = round(v) if field == "word_bigram_perplexity" else v
                except (ValueError, TypeError):
                    pass

        if len(values) != len(_EXPECTED_METRICS):
            return None

        return NgramSummary(author=xlsx_path.stem, **values)
    finally:
        wb.close()


def read_ngram_summaries(input_dir: Path) -> list[NgramSummary]:
    """Read all mega Excel files from a directory tree and return sorted summaries.

    Walks subdirectories (e.g. decade folders) to find all .xlsx files.
    """
    summaries: list[NgramSummary] = []
    for xlsx_path in sorted(input_dir.rglob("*.xlsx")):
        if xlsx_path.name.startswith("~$"):
            continue
        result = read_ngram_summary(xlsx_path)
        if result is not None:
            summaries.append(result)
    return sorted(summaries, key=lambda s: s.author.lower())


# ---------------------------------------------------------------------------
# Read – Prosody
# ---------------------------------------------------------------------------
_EXPECTED_PROSODY = {
    "Mean Syllables/Word": "mean_syllables_per_word",
    "Syllable Std Dev": "syllable_std_dev",
    "Polysyllabic Ratio (3+)": "polysyllabic_ratio",
    "Monosyllabic Ratio": "monosyllabic_ratio",
    "Rhythmic Regularity": "rhythmic_regularity",
    "Alliteration Density": "alliteration_density",
    "Assonance Density": "assonance_density",
    "Consonance Density": "consonance_density",
    "Sentence Rhythm Score": "sentence_rhythm_score",
    "Avg Consonant Cluster": "avg_consonant_cluster",
}


def read_prosody_summary(xlsx_path: Path) -> ProsodySummary | None:
    """Read Prosody stats from a single mega Excel file.

    Returns ``None`` if the Prosody sheet is missing or incomplete.
    """
    try:
        wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    except Exception:
        # Skip invalid/corrupted Excel files
        return None
    try:
        if "Prosody" not in wb.sheetnames:
            return None
        ws = wb["Prosody"]

        values: dict[str, float] = {}
        for row in ws.iter_rows(min_row=1, max_row=20, min_col=1, max_col=2):
            label = row[0].value
            val = row[1].value if len(row) > 1 else None
            if label and str(label).strip() in _EXPECTED_PROSODY and val is not None:
                try:
                    field = _EXPECTED_PROSODY[str(label).strip()]
                    values[field] = float(val)
                except (ValueError, TypeError):
                    pass

        if len(values) != len(_EXPECTED_PROSODY):
            return None

        return ProsodySummary(author=xlsx_path.stem, **values)
    finally:
        wb.close()


def read_prosody_summaries(input_dir: Path) -> list[ProsodySummary]:
    """Read Prosody sheets from all mega Excel files in a directory tree."""
    summaries: list[ProsodySummary] = []
    for xlsx_path in sorted(input_dir.rglob("*.xlsx")):
        if xlsx_path.name.startswith("~$"):
            continue
        result = read_prosody_summary(xlsx_path)
        if result is not None:
            summaries.append(result)
    return sorted(summaries, key=lambda s: s.author.lower())


# ---------------------------------------------------------------------------
# Read – Syllable Distribution
# ---------------------------------------------------------------------------
def read_syllable_summary(xlsx_path: Path) -> SyllableSummary | None:
    """Read syllable-count totals from the Prosody sheet's distribution table.

    The mega Prosody sheet contains a "Syllable Distribution" section with
    rows like::

        Syllables | Total | Unique
        1         | 54321 | 890
        2         | 23456 | 567
        ...
        12+       | 3     | 2

    We extract the "Total" column (column B) keyed by the bucket label
    (column A).  Returns ``None`` if no distribution data is found.

    Note:
        Backwards-compatible: also accepts the older "8+" overflow bucket
        from mega files generated before the extension to 12+.
    """
    try:
        wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    except Exception:
        # Skip invalid/corrupted Excel files
        return None
    try:
        if "Prosody" not in wb.sheetnames:
            return None
        ws = wb["Prosody"]

        # Scan for the "Syllables" sub-header row, then read data rows below it
        in_distribution = False
        totals: dict[str, int] = {}
        for row in ws.iter_rows(min_col=1, max_col=3):
            label = str(row[0].value).strip() if row[0].value is not None else ""

            if label == "Syllables":
                in_distribution = True
                continue

            if in_distribution:
                if label in {str(b) for b in range(1, 12)} | {"12+", "8+"}:
                    val = row[1].value if len(row) > 1 else None
                    try:
                        totals[label] = int(val) if val is not None else 0
                    except (ValueError, TypeError):
                        totals[label] = 0
                elif label and label not in {"", "None"}:
                    # Past the distribution section
                    break

        if not totals:
            return None

        return SyllableSummary(author=xlsx_path.stem, totals=totals)
    finally:
        wb.close()


def read_syllable_summaries(input_dir: Path) -> list[SyllableSummary]:
    """Read syllable distributions from all mega Excel files in a directory tree."""
    summaries: list[SyllableSummary] = []
    for xlsx_path in sorted(input_dir.rglob("*.xlsx")):
        if xlsx_path.name.startswith("~$"):
            continue
        result = read_syllable_summary(xlsx_path)
        if result is not None:
            summaries.append(result)
    return sorted(summaries, key=lambda s: s.author.lower())


# ---------------------------------------------------------------------------
# Read – Syllable Words (8–12+)
# ---------------------------------------------------------------------------
def read_syllable_words(input_dir: Path) -> dict[str, dict[str, int]]:
    """Read individual syllable words from Syllables tab for buckets 8–12+.

    Returns a dict mapping bucket -> {word: count_of_authors_with_word}.
    """
    HIGH_BUCKETS = ["8", "9", "10", "11", "12+"]
    word_counts: dict[str, dict[str, int]] = {b: {} for b in HIGH_BUCKETS}

    for xlsx_path in sorted(input_dir.rglob("*.xlsx")):
        if xlsx_path.name.startswith("~$"):
            continue
        try:
            wb = load_workbook(xlsx_path, read_only=True, data_only=True)
        except Exception:
            continue
        try:
            if "Syllables" not in wb.sheetnames:
                continue
            ws = wb["Syllables"]

            # Header row maps column index to bucket
            col_to_bucket: dict[int, str] = {}
            for col_idx, cell in enumerate(ws[1], start=1):
                if cell.value and str(cell.value) in HIGH_BUCKETS:
                    col_to_bucket[col_idx] = str(cell.value)

            if not col_to_bucket:
                continue

            # Read words from each bucket column
            # Use iter_rows with max_row limit to avoid infinite loop
            empty_row_count = 0
            for row in ws.iter_rows(min_row=2, max_row=10000, values_only=True):
                # Check if all bucket columns are empty
                has_content = False
                for col_idx, bucket in col_to_bucket.items():
                    cell_value = row[col_idx - 1] if col_idx <= len(row) else None
                    if cell_value and isinstance(cell_value, str) and cell_value.strip():
                        word = cell_value.strip().lower()
                        word_counts[bucket][word] = word_counts[bucket].get(word, 0) + 1
                        has_content = True

                # Stop if we hit 10 consecutive empty rows
                if not has_content:
                    empty_row_count += 1
                    if empty_row_count >= 10:
                        break
                else:
                    empty_row_count = 0
        finally:
            wb.close()

    return word_counts


# ---------------------------------------------------------------------------
# Read – Character Metrics
# ---------------------------------------------------------------------------
_EXPECTED_CHARACTER = {
    "Avg Word Length": "avg_word_length",
    "Avg Sentence Length (chars)": "avg_sentence_length_chars",
    "Punctuation Density": "punctuation_density",
    "Punctuation Variety": "punctuation_variety",
    "Vowel:Consonant Ratio": "vowel_consonant_ratio",
    "Digit Ratio": "digit_ratio",
    "Uppercase Ratio": "uppercase_ratio",
    "Whitespace Ratio": "whitespace_ratio",
}


def read_character_summary(xlsx_path: Path) -> CharacterSummary | None:
    """Read character-level metrics from a single mega Excel file.

    The Character sheet uses a two-column Metric/Value layout.
    Returns ``None`` if the sheet is missing or incomplete.
    """
    try:
        wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    except Exception:
        # Skip invalid/corrupted Excel files
        return None
    try:
        if "Character" not in wb.sheetnames:
            return None
        ws = wb["Character"]

        values: dict[str, float] = {}
        for row in ws.iter_rows(min_row=1, max_row=20, min_col=1, max_col=2):
            label = row[0].value
            val = row[1].value if len(row) > 1 else None
            if label and str(label).strip() in _EXPECTED_CHARACTER and val is not None:
                try:
                    field = _EXPECTED_CHARACTER[str(label).strip()]
                    values[field] = float(val)
                except (ValueError, TypeError):
                    pass

        if len(values) != len(_EXPECTED_CHARACTER):
            return None

        return CharacterSummary(author=xlsx_path.stem, **values)
    finally:
        wb.close()


def read_character_summaries(input_dir: Path) -> list[CharacterSummary]:
    """Read Character sheets from all mega Excel files in a directory tree."""
    summaries: list[CharacterSummary] = []
    for xlsx_path in sorted(input_dir.rglob("*.xlsx")):
        if xlsx_path.name.startswith("~$"):
            continue
        result = read_character_summary(xlsx_path)
        if result is not None:
            summaries.append(result)
    return sorted(summaries, key=lambda s: s.author.lower())


# ---------------------------------------------------------------------------
# Read – Lexical Diversity
# ---------------------------------------------------------------------------
_EXPECTED_LEXICAL = {
    "TTR": "ttr",
    "Root TTR (Guiraud)": "root_ttr",
    "Log TTR (Herdan)": "log_ttr",
    "STTR": "sttr",
    "MTLD (forward)": "mtld_forward",
    "MTLD (backward)": "mtld_backward",
    "MTLD (average)": "mtld_average",
    "Yule's K": "yule_k",
    "Yule's I": "yule_i",
    "Hapax Ratio": "hapax_ratio",
    "Dis-Hapax Ratio": "dis_hapax_ratio",
    "Sichel's S": "sichel_s",
    "Honore's R": "honore_r",
    "VocD-D": "vocd_d",
    "MATTR": "mattr",
    "HDD": "hdd",
    "MSTTR": "msttr",
}


def read_lexical_diversity_summary(xlsx_path: Path) -> LexicalDiversitySummary | None:
    """Read lexical diversity metrics from a single mega Excel file.

    The Lexical Diversity sheet uses a multi-column layout: Metric, Value, Std Dev, Min, Max.
    We only extract the Value column. Returns ``None`` if the sheet is missing.
    """
    try:
        wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    except Exception:
        return None
    try:
        if "Lexical Diversity" not in wb.sheetnames:
            return None
        ws = wb["Lexical Diversity"]

        values: dict[str, float] = {}
        for row in ws.iter_rows(min_row=2, max_row=50, min_col=1, max_col=2):
            label = row[0].value
            val = row[1].value if len(row) > 1 else None
            if label and str(label).strip() in _EXPECTED_LEXICAL and val is not None:
                try:
                    field = _EXPECTED_LEXICAL[str(label).strip()]
                    values[field] = float(val)
                except (ValueError, TypeError):
                    pass

        # Return summary even if some metrics are missing (they'll be None)
        if not values:
            return None

        return LexicalDiversitySummary(author=xlsx_path.stem, **values)
    finally:
        wb.close()


def read_lexical_diversity_summaries(input_dir: Path) -> list[LexicalDiversitySummary]:
    """Read Lexical Diversity sheets from all mega Excel files in a directory tree."""
    summaries: list[LexicalDiversitySummary] = []
    for xlsx_path in sorted(input_dir.rglob("*.xlsx")):
        if xlsx_path.name.startswith("~$"):
            continue
        result = read_lexical_diversity_summary(xlsx_path)
        if result is not None:
            summaries.append(result)
    return sorted(summaries, key=lambda s: s.author.lower())


# ---------------------------------------------------------------------------
# Read – BNC Summary
# ---------------------------------------------------------------------------
_EXPECTED_BNC = {
    "Total Tokens": "total_tokens",
    "Unique Tokens": "unique_tokens",
    "Overused": "overused",
    "Underused": "underused",
    "Not in BNC": "not_in_bnc",
}


def read_bnc_summary(xlsx_path: Path) -> BncSummary | None:
    """Read BNC summary stats from a single mega Excel file.

    Returns ``None`` if the BNC Summary sheet is missing or incomplete.
    """
    try:
        wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    except Exception:
        # Skip invalid/corrupted Excel files
        return None
    try:
        if "BNC Summary" not in wb.sheetnames:
            return None
        ws = wb["BNC Summary"]

        values: dict[str, int] = {}
        percentages: dict[str, float] = {}

        for row in ws.iter_rows(min_row=1, max_row=10, min_col=1, max_col=3):
            label = row[0].value
            count_val = row[1].value if len(row) > 1 else None
            pct_val = row[2].value if len(row) > 2 else None

            if label and str(label).strip() in _EXPECTED_BNC:
                field = _EXPECTED_BNC[str(label).strip()]
                try:
                    # Parse count value
                    if count_val is not None:
                        values[field] = int(count_val)

                    # Parse percentage value (skip Total Tokens which has no %)
                    if pct_val is not None and field != "total_tokens":
                        pct_str = str(pct_val).strip()
                        # Remove % sign if present
                        if pct_str.endswith("%"):
                            pct_str = pct_str[:-1]
                        percentages[f"pct_{field}"] = float(pct_str)
                except (ValueError, TypeError):
                    pass

        # Check we have all required counts
        if len(values) != len(_EXPECTED_BNC):
            return None

        return BncSummary(
            author=xlsx_path.stem,
            total_tokens=values["total_tokens"],
            unique_tokens=values["unique_tokens"],
            overused=values["overused"],
            underused=values["underused"],
            not_in_bnc=values["not_in_bnc"],
            pct_unique_tokens=percentages.get("pct_unique_tokens", 0.0),
            pct_overused=percentages.get("pct_overused", 0.0),
            pct_underused=percentages.get("pct_underused", 0.0),
            pct_not_in_bnc=percentages.get("pct_not_in_bnc", 0.0),
        )
    finally:
        wb.close()


def read_bnc_summaries(input_dir: Path) -> list[BncSummary]:
    """Read BNC Summary sheets from all mega Excel files in a directory tree."""
    summaries: list[BncSummary] = []
    for xlsx_path in sorted(input_dir.rglob("*.xlsx")):
        if xlsx_path.name.startswith("~$"):
            continue
        result = read_bnc_summary(xlsx_path)
        if result is not None:
            summaries.append(result)
    return sorted(summaries, key=lambda s: s.author.lower())


# ---------------------------------------------------------------------------
# Write
# ---------------------------------------------------------------------------
def _auto_width(ws: object, min_w: int = 15, max_w: int = 40) -> None:
    """Set reasonable column widths."""
    for col in ws.columns:
        letter = col[0].column_letter
        lengths = [len(str(c.value)) for c in col if c.value is not None]
        best = max(lengths) if lengths else min_w
        ws.column_dimensions[letter].width = min(max(best + 2, min_w), max_w)


def _style_sheet(ws: object) -> None:
    """Apply consistent header + data styling to a sheet."""
    for cell in ws[1]:
        cell.font = _FONT_HEADER
        cell.fill = _FILL_HEADER
        cell.alignment = _ALIGN

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = _ALIGN
        row[0].alignment = _ALIGN_LEFT

    ws.row_dimensions[1].height = 25
    for r in range(2, ws.max_row + 1):
        ws.row_dimensions[r].height = 20

    _auto_width(ws)


def write_mega_meta_excel(
    summaries: list[NgramSummary],
    output_path: Path,
    prosody_summaries: list[ProsodySummary] | None = None,
    syllable_summaries: list[SyllableSummary] | None = None,
    character_summaries: list[CharacterSummary] | None = None,
    syllable_words: dict[str, dict[str, int]] | None = None,
    lexical_diversity_summaries: list[LexicalDiversitySummary] | None = None,
    bnc_summaries: list[BncSummary] | None = None,
) -> None:
    """Write comparative mega meta-analysis to an Excel workbook.

    Creates sheets:
      - ``N-grams``: entropy and perplexity per author (authors as rows)
      - ``Prosody``: rhythm and prosody metrics per author (if available)
      - ``Syllables-Tot``: raw syllable counts per bucket (if available)
      - ``Syllables-Pct``: syllable percentages per bucket (if available)
      - ``Syllables-Txt``: word lists with counts for buckets 8–12+ (if available)
      - ``Character``: character-level metrics per author (if available)
      - ``Lexical Diversity``: lexical diversity metrics per author (if available)
      - ``BNC Summary Tot``: BNC token counts per author (if available)
      - ``BNC Summary Pct``: BNC percentages per author (if available)
    """
    wb = Workbook()
    wb.remove(wb.active)

    # ── N-grams tab ───────────────────────────────────────────────────────
    ws = wb.create_sheet("N-grams")
    ws.append(["Author"] + NGRAM_METRIC_LABELS)

    for s in summaries:
        ws.append([
            s.author,
            s.char_bigram_entropy,
            s.char_bigram_perplexity,
            s.word_bigram_entropy,
            s.word_bigram_perplexity,
        ])

    _style_sheet(ws)

    # Number formats: per-column (col 5 = Word Bigram Perplexity is integer)
    _NGRAM_COL_FORMATS = {2: "#,##0.0000", 3: "#,##0.0000", 4: "#,##0.0000", 5: "#,##0"}
    for col_idx, fmt in _NGRAM_COL_FORMATS.items():
        for row_idx in range(2, ws.max_row + 1):
            ws.cell(row=row_idx, column=col_idx).number_format = fmt

    # Conditional formatting: column-by-column heatmap (white=low, green=high)
    # Skip columns where all values are equal
    for col_idx in range(2, len(NGRAM_METRIC_LABELS) + 2):
        # Check if all values in column are equal
        values = [ws.cell(row=r, column=col_idx).value for r in range(2, ws.max_row + 1)]
        if len(set(values)) > 1:  # More than one unique value
            col_letter = get_column_letter(col_idx)
            cell_range = f"{col_letter}2:{col_letter}{ws.max_row}"
            rule = ColorScaleRule(
                start_type="min", start_color="FFFFFF",
                end_type="max", end_color="63BE7B"
            )
            ws.conditional_formatting.add(cell_range, rule)

    # ── Prosody tab ───────────────────────────────────────────────────────
    if prosody_summaries:
        ws_p = wb.create_sheet("Prosody")
        ws_p.append(["Author"] + PROSODY_METRIC_LABELS)

        for p in prosody_summaries:
            ws_p.append([
                p.author,
                p.mean_syllables_per_word,
                p.syllable_std_dev,
                p.polysyllabic_ratio,
                p.monosyllabic_ratio,
                p.rhythmic_regularity,
                p.alliteration_density,
                p.assonance_density,
                p.consonance_density,
                p.sentence_rhythm_score,
                p.avg_consonant_cluster,
            ])

        _style_sheet(ws_p)

        # Number formats: all prosody metrics are floats
        for col_idx in range(2, len(PROSODY_METRIC_LABELS) + 2):
            for row_idx in range(2, ws_p.max_row + 1):
                ws_p.cell(row=row_idx, column=col_idx).number_format = "#,##0.0000"

        # Conditional formatting: column-by-column heatmap (white=low, green=high)
        # Skip columns where all values are equal
        for col_idx in range(2, len(PROSODY_METRIC_LABELS) + 2):
            # Check if all values in column are equal
            values = [ws_p.cell(row=r, column=col_idx).value for r in range(2, ws_p.max_row + 1)]
            if len(set(values)) > 1:  # More than one unique value
                col_letter = get_column_letter(col_idx)
                cell_range = f"{col_letter}2:{col_letter}{ws_p.max_row}"
                rule = ColorScaleRule(
                    start_type="min", start_color="FFFFFF",
                    end_type="max", end_color="63BE7B"
                )
                ws_p.conditional_formatting.add(cell_range, rule)

    # ── Syllables-Tot tab ──────────────────────────────────────────────────
    # Authors as rows, syllable buckets (1–11, 12+) as columns, raw totals
    if syllable_summaries:
        ws_s_tot = wb.create_sheet("Syllables-Tot")
        ws_s_tot.append(["Author"] + SYLLABLE_BUCKETS)

        for s in syllable_summaries:
            ws_s_tot.append(
                [s.author] + [s.totals.get(b, 0) for b in SYLLABLE_BUCKETS]
            )

        _style_sheet(ws_s_tot)

        # Number formats: integer with thousand separators
        for col_idx in range(2, len(SYLLABLE_BUCKETS) + 2):
            for row_idx in range(2, ws_s_tot.max_row + 1):
                ws_s_tot.cell(row=row_idx, column=col_idx).number_format = "#,##0"

        # Conditional formatting: column-by-column heatmap (white=low, green=high)
        # Skip columns where all values are equal
        for col_idx in range(2, len(SYLLABLE_BUCKETS) + 2):
            # Check if all values in column are equal
            values = [ws_s_tot.cell(row=r, column=col_idx).value for r in range(2, ws_s_tot.max_row + 1)]
            if len(set(values)) > 1:  # More than one unique value
                col_letter = get_column_letter(col_idx)
                cell_range = f"{col_letter}2:{col_letter}{ws_s_tot.max_row}"
                rule = ColorScaleRule(
                    start_type="min", start_color="FFFFFF",
                    end_type="max", end_color="63BE7B"
                )
                ws_s_tot.conditional_formatting.add(cell_range, rule)

    # ── Syllables-Pct tab ──────────────────────────────────────────────────
    # Authors as rows, syllable buckets (1–11, 12+) as columns, percentages
    if syllable_summaries:
        ws_s_pct = wb.create_sheet("Syllables-Pct")
        ws_s_pct.append(["Author"] + SYLLABLE_BUCKETS)

        for s in syllable_summaries:
            total_words = sum(s.totals.get(b, 0) for b in SYLLABLE_BUCKETS)
            if total_words > 0:
                percentages = [
                    (s.totals.get(b, 0) / total_words * 100) for b in SYLLABLE_BUCKETS
                ]
            else:
                percentages = [0.0] * len(SYLLABLE_BUCKETS)
            ws_s_pct.append([s.author] + percentages)

        _style_sheet(ws_s_pct)

        # Number formats: percentage with 6 decimals
        for col_idx in range(2, len(SYLLABLE_BUCKETS) + 2):
            for row_idx in range(2, ws_s_pct.max_row + 1):
                ws_s_pct.cell(row=row_idx, column=col_idx).number_format = "0.000000"

        # Conditional formatting: column-by-column heatmap (white=low, green=high)
        # Skip columns where all values are equal
        for col_idx in range(2, len(SYLLABLE_BUCKETS) + 2):
            # Check if all values in column are equal
            values = [ws_s_pct.cell(row=r, column=col_idx).value for r in range(2, ws_s_pct.max_row + 1)]
            if len(set(values)) > 1:  # More than one unique value
                col_letter = get_column_letter(col_idx)
                cell_range = f"{col_letter}2:{col_letter}{ws_s_pct.max_row}"
                rule = ColorScaleRule(
                    start_type="min", start_color="FFFFFF",
                    end_type="max", end_color="63BE7B"
                )
                ws_s_pct.conditional_formatting.add(cell_range, rule)

    # ── Syllables-Txt tab ──────────────────────────────────────────────────
    # Word lists with counts for syllable buckets 8–12+
    if syllable_words:
        HIGH_BUCKETS = ["8", "9", "10", "11", "12+"]
        ws_txt = wb.create_sheet("Syllables-Txt")

        # Header row: bucket, count, bucket, count, ...
        header = []
        for bucket in HIGH_BUCKETS:
            header.extend([bucket, "Count"])
        ws_txt.append(header)

        # Sort words by count (descending) for each bucket
        sorted_words = {}
        for bucket in HIGH_BUCKETS:
            words_in_bucket = syllable_words.get(bucket, {})
            sorted_words[bucket] = sorted(
                words_in_bucket.items(), key=lambda x: x[1], reverse=True
            )

        # Find max row count across all buckets
        max_rows = max((len(sorted_words[b]) for b in HIGH_BUCKETS), default=0)

        # Write data rows
        for row_idx in range(max_rows):
            row = []
            for bucket in HIGH_BUCKETS:
                if row_idx < len(sorted_words[bucket]):
                    word, count = sorted_words[bucket][row_idx]
                    row.extend([word, count])
                else:
                    row.extend([None, None])
            ws_txt.append(row)

        _style_sheet(ws_txt)
        _auto_width(ws_txt, min_w=20)

    # ── Character tab ──────────────────────────────────────────────────────
    if character_summaries:
        ws_c = wb.create_sheet("Character")
        ws_c.append(["Author"] + CHARACTER_METRIC_LABELS)

        for c in character_summaries:
            ws_c.append([
                c.author,
                c.avg_word_length,
                c.avg_sentence_length_chars,
                c.punctuation_density,
                c.punctuation_variety,
                c.vowel_consonant_ratio,
                c.digit_ratio,
                c.uppercase_ratio,
                c.whitespace_ratio,
            ])

        _style_sheet(ws_c)

        for col_idx in range(2, len(CHARACTER_METRIC_LABELS) + 2):
            for row_idx in range(2, ws_c.max_row + 1):
                ws_c.cell(row=row_idx, column=col_idx).number_format = "#,##0.0000"

        # Conditional formatting: column-by-column heatmap (white=low, green=high)
        # Skip columns where all values are equal
        for col_idx in range(2, len(CHARACTER_METRIC_LABELS) + 2):
            # Check if all values in column are equal
            values = [ws_c.cell(row=r, column=col_idx).value for r in range(2, ws_c.max_row + 1)]
            if len(set(values)) > 1:  # More than one unique value
                col_letter = get_column_letter(col_idx)
                cell_range = f"{col_letter}2:{col_letter}{ws_c.max_row}"
                rule = ColorScaleRule(
                    start_type="min", start_color="FFFFFF",
                    end_type="max", end_color="63BE7B"
                )
                ws_c.conditional_formatting.add(cell_range, rule)

    # ── Lexical Diversity tab ──────────────────────────────────────────────
    if lexical_diversity_summaries:
        ws_ld = wb.create_sheet("Lexical Diversity")
        ws_ld.append(["Author"] + LEXICAL_METRIC_LABELS)

        for ld in lexical_diversity_summaries:
            ws_ld.append([
                ld.author,
                ld.ttr,
                ld.root_ttr,
                ld.log_ttr,
                ld.sttr,
                ld.mtld_forward,
                ld.mtld_backward,
                ld.mtld_average,
                ld.yule_k,
                ld.yule_i,
                ld.hapax_ratio,
                ld.dis_hapax_ratio,
                ld.sichel_s,
                ld.honore_r,
                ld.vocd_d,
                ld.mattr,
                ld.hdd,
                ld.msttr,
            ])

        _style_sheet(ws_ld)

        for col_idx in range(2, len(LEXICAL_METRIC_LABELS) + 2):
            for row_idx in range(2, ws_ld.max_row + 1):
                ws_ld.cell(row=row_idx, column=col_idx).number_format = "#,##0.0000"

        # Conditional formatting: column-by-column heatmap (white=low, green=high)
        # Skip columns where all values are equal
        for col_idx in range(2, len(LEXICAL_METRIC_LABELS) + 2):
            # Check if all values in column are equal
            values = [ws_ld.cell(row=r, column=col_idx).value for r in range(2, ws_ld.max_row + 1)]
            if len(set(values)) > 1:  # More than one unique value
                col_letter = get_column_letter(col_idx)
                cell_range = f"{col_letter}2:{col_letter}{ws_ld.max_row}"
                rule = ColorScaleRule(
                    start_type="min", start_color="FFFFFF",
                    end_type="max", end_color="63BE7B"
                )
                ws_ld.conditional_formatting.add(cell_range, rule)

    # ── BNC Summary Tot tab ────────────────────────────────────────────────
    if bnc_summaries:
        ws_bnc_tot = wb.create_sheet("BNC Summary Tot")
        ws_bnc_tot.append(["Author"] + BNC_METRIC_LABELS)

        for bnc in bnc_summaries:
            ws_bnc_tot.append([
                bnc.author,
                bnc.total_tokens,
                bnc.unique_tokens,
                bnc.overused,
                bnc.underused,
                bnc.not_in_bnc,
            ])

        _style_sheet(ws_bnc_tot)

        # Number formats: integer with thousand separators
        for col_idx in range(2, len(BNC_METRIC_LABELS) + 2):
            for row_idx in range(2, ws_bnc_tot.max_row + 1):
                ws_bnc_tot.cell(row=row_idx, column=col_idx).number_format = "#,##0"

        # Conditional formatting: column-by-column heatmap (white=low, green=high)
        # Skip columns where all values are equal
        for col_idx in range(2, len(BNC_METRIC_LABELS) + 2):
            # Check if all values in column are equal
            values = [ws_bnc_tot.cell(row=r, column=col_idx).value for r in range(2, ws_bnc_tot.max_row + 1)]
            if len(set(values)) > 1:  # More than one unique value
                col_letter = get_column_letter(col_idx)
                cell_range = f"{col_letter}2:{col_letter}{ws_bnc_tot.max_row}"
                rule = ColorScaleRule(
                    start_type="min", start_color="FFFFFF",
                    end_type="max", end_color="63BE7B"
                )
                ws_bnc_tot.conditional_formatting.add(cell_range, rule)

    # ── BNC Summary Pct tab ────────────────────────────────────────────────
    if bnc_summaries:
        ws_bnc_pct = wb.create_sheet("BNC Summary Pct")
        ws_bnc_pct.append(["Author"] + BNC_PCT_METRIC_LABELS)

        for bnc in bnc_summaries:
            ws_bnc_pct.append([
                bnc.author,
                bnc.pct_unique_tokens,
                bnc.pct_overused,
                bnc.pct_underused,
                bnc.pct_not_in_bnc,
            ])

        _style_sheet(ws_bnc_pct)

        # Number formats: percentage with 2 decimals
        for col_idx in range(2, len(BNC_PCT_METRIC_LABELS) + 2):
            for row_idx in range(2, ws_bnc_pct.max_row + 1):
                ws_bnc_pct.cell(row=row_idx, column=col_idx).number_format = "0.00"

        # Conditional formatting: column-by-column heatmap (white=low, green=high)
        # Skip columns where all values are equal
        for col_idx in range(2, len(BNC_PCT_METRIC_LABELS) + 2):
            # Check if all values in column are equal
            values = [ws_bnc_pct.cell(row=r, column=col_idx).value for r in range(2, ws_bnc_pct.max_row + 1)]
            if len(set(values)) > 1:  # More than one unique value
                col_letter = get_column_letter(col_idx)
                cell_range = f"{col_letter}2:{col_letter}{ws_bnc_pct.max_row}"
                rule = ColorScaleRule(
                    start_type="min", start_color="FFFFFF",
                    end_type="max", end_color="63BE7B"
                )
                ws_bnc_pct.conditional_formatting.add(cell_range, rule)

    wb.save(output_path)
