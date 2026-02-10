"""Mega analysis: comprehensive stylometric fingerprint in a single Excel workbook.

Runs every applicable single-text analysis in pystylometry and writes results
to a multi-tab Excel file.  Optional-dependency modules (readability, syntactic,
prosody) are attempted and gracefully skipped when unavailable.

Analyses that require *two* texts (Burrows' Delta, Zeta, Kilgarriff comparison,
NCD, vocabulary overlap) are excluded — mega is a single-corpus tool.

Usage (via CLI):
    mega --input-file manuscript.txt --output-file /tmp/out.xlsx
    mega --input-dir ~/ebooks/Author --output-file /tmp/author.xlsx
"""

from __future__ import annotations

import math
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from rich.console import Console

# ---------------------------------------------------------------------------
# Shared Excel styles (match existing bnc_frequency_cli patterns)
# ---------------------------------------------------------------------------
_ALIGN = Alignment(horizontal="center", vertical="center")
_ALIGN_LEFT = Alignment(horizontal="left", vertical="center")
_FONT_HEADER = Font(bold=True, size=14)
_FILL_HEADER = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
_FILL_SECTION = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
_FONT_SECTION = Font(bold=True, size=12)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def _fmt(val: Any) -> Any:
    """Format a value for Excel: round floats, stringify NaN."""
    if isinstance(val, float):
        if math.isnan(val) or math.isinf(val):
            return "N/A"
        return round(val, 4)
    return val


def _style_header(ws: Any) -> None:
    """Apply header styling to the first row of a worksheet."""
    for cell in ws[1]:
        cell.font = _FONT_HEADER
        cell.fill = _FILL_HEADER
        cell.alignment = _ALIGN


def _auto_width(ws: Any, min_w: int = 15, max_w: int = 40) -> None:
    """Set reasonable column widths."""
    for col in ws.columns:
        letter = col[0].column_letter
        lengths = []
        for cell in col:
            if cell.value is not None:
                lengths.append(len(str(cell.value)))
        best = max(lengths) if lengths else min_w
        ws.column_dimensions[letter].width = min(max(best + 2, min_w), max_w)


def _align_all(ws: Any) -> None:
    """Center-align all cells."""
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = _ALIGN


def _set_row_heights(ws: Any, header_h: int = 25, data_h: int = 20) -> None:
    """Set uniform row heights."""
    for r in range(1, ws.max_row + 1):
        ws.row_dimensions[r].height = data_h
    ws.row_dimensions[1].height = header_h


def _skip_tab(ws: Any, package: str, extra: str) -> None:
    """Write a single-row skip message into a worksheet."""
    ws.append(["Skipped", f"Requires {package}. Install with: pip install pystylometry[{extra}]"])
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 60


# ---------------------------------------------------------------------------
# Analysis runner
# ---------------------------------------------------------------------------
def run_mega(text: str, console: Console) -> dict[str, Any]:
    """Run all applicable analyses and return a results dict.

    Keys correspond to tab names. A missing key means the module was skipped.
    The special key ``_skipped`` lists (tab, reason) tuples.
    """
    results: dict[str, Any] = {}
    skipped: list[tuple[str, str]] = []

    # -- Lexical (always available) -----------------------------------------
    console.print("  [dim]Running[/dim] lexical diversity...", end="")
    from pystylometry.lexical import (
        compute_function_words,
        compute_hapax_ratios,
        compute_hdd,
        compute_mattr,
        compute_msttr,
        compute_mtld,
        compute_ttr,
        compute_vocd_d,
        compute_yule,
    )

    results["lexical"] = {
        "ttr": compute_ttr(text),
        "mtld": compute_mtld(text),
        "yule": compute_yule(text),
        "hapax": compute_hapax_ratios(text),
        "vocd_d": compute_vocd_d(text),
        "mattr": compute_mattr(text),
        "hdd": compute_hdd(text),
        "msttr": compute_msttr(text),
    }
    results["function_words"] = compute_function_words(text)
    console.print(" [green]done[/green]")

    # -- BNC Frequency (requires bnc-lookup) --------------------------------
    try:
        console.print("  [dim]Running[/dim] BNC frequency...", end="")
        from pystylometry.lexical import compute_bnc_frequency

        results["bnc"] = compute_bnc_frequency(text)
        console.print(" [green]done[/green]")
    except (ImportError, Exception) as exc:
        skipped.append(("BNC Frequency", str(exc)))
        console.print(f" [yellow]skipped[/yellow] ({exc})")

    # -- Word Frequency Sophistication (requires bnc-lookup) ----------------
    try:
        console.print("  [dim]Running[/dim] word frequency sophistication...", end="")
        from pystylometry.lexical import compute_word_frequency_sophistication

        results["freq_soph"] = compute_word_frequency_sophistication(text)
        console.print(" [green]done[/green]")
    except (ImportError, Exception) as exc:
        skipped.append(("Frequency Sophistication", str(exc)))
        console.print(f" [yellow]skipped[/yellow] ({exc})")

    # -- Repetition (requires bnc-lookup) -----------------------------------
    try:
        console.print("  [dim]Running[/dim] repetition / slop detection...", end="")
        from pystylometry.lexical import compute_repetitive_ngrams, compute_repetitive_unigrams

        results["rep_uni"] = compute_repetitive_unigrams(text)
        results["rep_ng"] = compute_repetitive_ngrams(text)
        console.print(" [green]done[/green]")
    except (ImportError, Exception) as exc:
        skipped.append(("Repetition", str(exc)))
        console.print(f" [yellow]skipped[/yellow] ({exc})")

    # -- Character metrics (always available) --------------------------------
    console.print("  [dim]Running[/dim] character metrics...", end="")
    from pystylometry.character import compute_character_metrics

    results["character"] = compute_character_metrics(text)
    console.print(" [green]done[/green]")

    # -- N-grams (always available) -----------------------------------------
    console.print("  [dim]Running[/dim] n-gram entropy...", end="")
    from pystylometry.ngrams import (
        compute_character_bigram_entropy,
        compute_word_bigram_entropy,
    )

    results["ngrams"] = {
        "char_bi": compute_character_bigram_entropy(text),
        "word_bi": compute_word_bigram_entropy(text),
    }

    # Extended n-grams (trigrams, 4-grams, skipgrams) — may need spaCy for POS
    try:
        from pystylometry.ngrams import compute_extended_ngrams

        results["ngrams"]["extended"] = compute_extended_ngrams(text, top_n=15, include_pos_ngrams=False)
    except (ImportError, Exception):
        pass  # POS ngrams need spaCy; skip silently, entropy tabs still populated

    console.print(" [green]done[/green]")

    # -- Stylistic markers (always available) --------------------------------
    console.print("  [dim]Running[/dim] stylistic markers...", end="")
    from pystylometry.stylistic import compute_stylistic_markers

    results["markers"] = compute_stylistic_markers(text)
    console.print(" [green]done[/green]")

    # -- Genre & Register (always available) ---------------------------------
    console.print("  [dim]Running[/dim] genre & register...", end="")
    from pystylometry.stylistic import compute_genre_register

    results["genre"] = compute_genre_register(text)
    console.print(" [green]done[/green]")

    # -- Dialect (always available) ------------------------------------------
    console.print("  [dim]Running[/dim] dialect detection...", end="")
    from pystylometry.dialect import compute_dialect

    results["dialect"] = compute_dialect(text)
    console.print(" [green]done[/green]")

    # -- Drift (always available) --------------------------------------------
    console.print("  [dim]Running[/dim] style drift...", end="")
    from pystylometry.consistency import compute_kilgarriff_drift

    results["drift"] = compute_kilgarriff_drift(text)
    console.print(" [green]done[/green]")

    # -- Readability (requires pronouncing) ----------------------------------
    try:
        console.print("  [dim]Running[/dim] readability indices...", end="")
        from pystylometry.readability import (
            compute_ari,
            compute_coleman_liau,
            compute_dale_chall,
            compute_flesch,
            compute_forcast,
            compute_fry,
            compute_gunning_fog,
            compute_linsear_write,
            compute_powers_sumner_kearl,
            compute_smog,
        )

        results["readability"] = {
            "flesch": compute_flesch(text),
            "smog": compute_smog(text),
            "ari": compute_ari(text),
            "coleman_liau": compute_coleman_liau(text),
            "gunning_fog": compute_gunning_fog(text),
            "dale_chall": compute_dale_chall(text),
            "linsear": compute_linsear_write(text),
            "fry": compute_fry(text),
            "forcast": compute_forcast(text),
            "psk": compute_powers_sumner_kearl(text),
        }
        console.print(" [green]done[/green]")
    except ImportError as exc:
        skipped.append(("Readability", str(exc)))
        console.print(f" [yellow]skipped[/yellow] ({exc})")

    # -- Prosody (requires pronouncing) --------------------------------------
    try:
        console.print("  [dim]Running[/dim] rhythm & prosody...", end="")
        from pystylometry.prosody import compute_rhythm_prosody

        results["prosody"] = compute_rhythm_prosody(text)
        console.print(" [green]done[/green]")
    except ImportError as exc:
        skipped.append(("Prosody", str(exc)))
        console.print(f" [yellow]skipped[/yellow] ({exc})")

    # -- Syntactic (requires spaCy) ------------------------------------------
    # spaCy's default max_length is 1M chars; skip proactively for large texts
    _SPACY_MAX = 1_000_000
    if len(text) > _SPACY_MAX:
        _skip_reason = f"Text too large for spaCy ({len(text):,} chars > {_SPACY_MAX:,} limit)"
        console.print(f"  [dim]Running[/dim] syntactic analysis... [yellow]skipped[/yellow] ({_skip_reason})")
        skipped.append(("Syntactic", _skip_reason))
        console.print(f"  [dim]Running[/dim] cohesion & coherence... [yellow]skipped[/yellow] ({_skip_reason})")
        skipped.append(("Cohesion", _skip_reason))
    else:
        try:
            console.print("  [dim]Running[/dim] syntactic analysis...", end="")
            from pystylometry.syntactic import (
                compute_advanced_syntactic,
                compute_pos_ratios,
                compute_sentence_stats,
                compute_sentence_types,
            )

            results["syntactic"] = {
                "pos": compute_pos_ratios(text),
                "sent_stats": compute_sentence_stats(text),
                "sent_types": compute_sentence_types(text),
                "advanced": compute_advanced_syntactic(text),
            }
            console.print(" [green]done[/green]")
        except (ImportError, Exception) as exc:
            skipped.append(("Syntactic", str(exc)))
            console.print(f" [yellow]skipped[/yellow] ({exc})")

        # -- Cohesion & Coherence (requires spaCy) ---------------------------
        try:
            console.print("  [dim]Running[/dim] cohesion & coherence...", end="")
            from pystylometry.stylistic import compute_cohesion_coherence

            results["cohesion"] = compute_cohesion_coherence(text)
            console.print(" [green]done[/green]")
        except (ImportError, Exception) as exc:
            skipped.append(("Cohesion", str(exc)))
            console.print(f" [yellow]skipped[/yellow] ({exc})")

    results["_skipped"] = skipped
    return results


# ---------------------------------------------------------------------------
# Excel writer
# ---------------------------------------------------------------------------
def write_mega_excel(results: dict[str, Any], output_path: Path) -> None:
    """Write all analysis results to a multi-tab Excel workbook."""
    wb = Workbook()
    wb.remove(wb.active)  # type: ignore[arg-type]

    # ── 1. Dashboard ───────────────────────────────────────────────────────
    ws = wb.create_sheet("Dashboard")
    ws.append(["Category", "Metric", "Value", "Description"])
    _style_header(ws)

    _dashboard_rows(ws, results)

    # Consistent decimal formatting (#.####) for the Value column (C)
    for row in ws.iter_rows(min_row=2, min_col=3, max_col=3):
        for cell in row:
            if isinstance(cell.value, float):
                cell.number_format = "0.0000"

    _auto_width(ws, min_w=18, max_w=50)
    _align_all(ws)
    _set_row_heights(ws)

    # ── 2. Lexical Diversity ───────────────────────────────────────────────
    ws = wb.create_sheet("Lexical Diversity")
    ws.append(["Metric", "Value", "Std Dev", "Min", "Max"])
    _style_header(ws)
    _lexical_tab(ws, results.get("lexical", {}))
    _auto_width(ws)
    _align_all(ws)
    _set_row_heights(ws)

    # ── 3. Readability ─────────────────────────────────────────────────────
    ws = wb.create_sheet("Readability")
    if "readability" in results:
        ws.append(["Index", "Score", "Std Dev"])
        _style_header(ws)
        _readability_tab(ws, results["readability"])
    else:
        _skip_tab(ws, "pronouncing", "readability")
    _auto_width(ws)
    _align_all(ws)
    _set_row_heights(ws)

    # ── 4. Sentences ───────────────────────────────────────────────────────
    ws = wb.create_sheet("Sentences")
    if "syntactic" in results:
        ws.append(["Metric", "Value", "Std Dev"])
        _style_header(ws)
        _sentences_tab(ws, results["syntactic"])
    else:
        _skip_tab(ws, "spacy", "syntactic")
    _auto_width(ws)
    _align_all(ws)
    _set_row_heights(ws)

    # ── 5. Function Words ─────────────────────────────────────────────────
    ws = wb.create_sheet("Function Words")
    ws.append(["Metric", "Value"])
    _style_header(ws)
    _function_words_tab(ws, results.get("function_words"))
    _auto_width(ws)
    _align_all(ws)
    _set_row_heights(ws)

    # ── 6. Style Markers ──────────────────────────────────────────────────
    ws = wb.create_sheet("Style Markers")
    ws.append(["Marker", "Value"])
    _style_header(ws)
    _style_markers_tab(ws, results.get("markers"))
    _auto_width(ws)
    _align_all(ws)
    _set_row_heights(ws)

    # ── 7. Character ──────────────────────────────────────────────────────
    ws = wb.create_sheet("Character")
    ws.append(["Metric", "Value", "Std Dev"])
    _style_header(ws)
    _character_tab(ws, results.get("character"))
    _auto_width(ws)
    _align_all(ws)
    _set_row_heights(ws)

    # ── 8. N-grams ────────────────────────────────────────────────────────
    ws = wb.create_sheet("N-grams")
    ws.append(["Metric", "Value"])
    _style_header(ws)
    _ngrams_tab(ws, results.get("ngrams", {}))
    _auto_width(ws)
    _align_all(ws)
    _set_row_heights(ws)

    # ── 9. Prosody ────────────────────────────────────────────────────────
    ws = wb.create_sheet("Prosody")
    if "prosody" in results:
        ws.append(["Metric", "Value"])
        _style_header(ws)
        _prosody_tab(ws, results["prosody"])
    else:
        _skip_tab(ws, "pronouncing", "readability")
    _auto_width(ws)
    _align_all(ws)
    _set_row_heights(ws)

    # ── 10. Cohesion ──────────────────────────────────────────────────────
    ws = wb.create_sheet("Cohesion")
    if "cohesion" in results:
        ws.append(["Metric", "Value"])
        _style_header(ws)
        _cohesion_tab(ws, results["cohesion"])
    else:
        _skip_tab(ws, "spacy", "syntactic")
    _auto_width(ws)
    _align_all(ws)
    _set_row_heights(ws)

    # ── 11. Genre & Register ──────────────────────────────────────────────
    ws = wb.create_sheet("Genre & Register")
    ws.append(["Metric", "Value"])
    _style_header(ws)
    _genre_tab(ws, results.get("genre"))
    _auto_width(ws)
    _align_all(ws)
    _set_row_heights(ws)

    # ── 12. Dialect ───────────────────────────────────────────────────────
    ws = wb.create_sheet("Dialect")
    ws.append(["Metric", "Value"])
    _style_header(ws)
    _dialect_tab(ws, results.get("dialect"))
    _auto_width(ws)
    _align_all(ws)
    _set_row_heights(ws)

    # ── 13. Drift ─────────────────────────────────────────────────────────
    ws = wb.create_sheet("Style Drift")
    ws.append(["Metric", "Value"])
    _style_header(ws)
    _drift_tab(ws, results.get("drift"))
    _auto_width(ws)
    _align_all(ws)
    _set_row_heights(ws)

    # ── 14. Repetition ────────────────────────────────────────────────────
    ws = wb.create_sheet("Repetition")
    if "rep_uni" in results:
        ws.append(["Word / N-gram", "Count", "Expected", "Repetition Score"])
        _style_header(ws)
        _repetition_tab(ws, results.get("rep_uni"), results.get("rep_ng"))
    else:
        _skip_tab(ws, "bnc-lookup", "lexical")
    _auto_width(ws)
    _align_all(ws)
    _set_row_heights(ws)

    # ── 15-17. BNC detail tabs ────────────────────────────────────────────
    if "bnc" in results:
        _bnc_detail_tabs(wb, results["bnc"])
    else:
        for name in ("BNC Overused", "BNC Underused", "BNC Not Found"):
            ws = wb.create_sheet(name)
            _skip_tab(ws, "bnc-lookup", "lexical")

    wb.save(output_path)


# ---------------------------------------------------------------------------
# Tab writers
# ---------------------------------------------------------------------------
def _section_row(ws: Any, label: str) -> None:
    """Insert a section divider row."""
    row_num = ws.max_row + 1
    c = ws.cell(row=row_num, column=1, value=label)
    c.font = _FONT_SECTION
    c.fill = _FILL_SECTION
    for col in range(2, ws.max_column + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.fill = _FILL_SECTION


def _dashboard_rows(ws: Any, results: dict[str, Any]) -> None:
    """Populate the Dashboard tab with one row per key metric."""

    def _r(cat: str, metric: str, val: Any, desc: str = "") -> None:
        ws.append([cat, metric, _fmt(val), desc])

    # -- Lexical --
    lex = results.get("lexical", {})
    if lex:
        ttr = lex.get("ttr")
        if ttr:
            _r("Lexical", "TTR", ttr.ttr, "Type-Token Ratio")
            _r("Lexical", "Root TTR", ttr.root_ttr, "Guiraud's Root TTR")
            _r("Lexical", "Log TTR", ttr.log_ttr, "Herdan's Log TTR")
            _r("Lexical", "STTR", ttr.sttr, "Standardized TTR (1000-token chunks)")
        mtld = lex.get("mtld")
        if mtld:
            _r("Lexical", "MTLD", mtld.mtld_average, "Measure of Textual Lexical Diversity")
        yule = lex.get("yule")
        if yule:
            _r("Lexical", "Yule's K", yule.yule_k, "Vocabulary repetitiveness (higher = more)")
            _r("Lexical", "Yule's I", yule.yule_i, "Vocabulary richness (higher = richer)")
        hapax = lex.get("hapax")
        if hapax:
            _r("Lexical", "Hapax Ratio", hapax.hapax_ratio, "Words appearing exactly once / total")
            _r("Lexical", "Sichel's S", hapax.sichel_s, "Dis-legomena / vocabulary size")
            _r("Lexical", "Honore's R", hapax.honore_r, "Hapax-based vocabulary richness")
        vocd = lex.get("vocd_d")
        if vocd:
            _r("Lexical", "VocD-D", vocd.d_parameter, "Vocabulary diversity (curve-fitting)")
        mattr = lex.get("mattr")
        if mattr:
            _r("Lexical", "MATTR", mattr.mattr_score, "Moving-Average TTR")
        hdd = lex.get("hdd")
        if hdd:
            _r("Lexical", "HDD", hdd.hdd_score, "Hypergeometric distribution D")
        msttr = lex.get("msttr")
        if msttr:
            _r("Lexical", "MSTTR", msttr.msttr_score, "Mean Segmental TTR")

    # -- Readability --
    rd = results.get("readability", {})
    if rd:
        fl = rd.get("flesch")
        if fl:
            _r("Readability", "Flesch Reading Ease", fl.reading_ease, "0-100 (higher = easier)")
        sm = rd.get("smog")
        if sm:
            _r("Readability", "SMOG Grade", sm.smog_index, "Grade level")
        ar = rd.get("ari")
        if ar:
            _r("Readability", "ARI", ar.ari_score, "Automated Readability Index")
        cl = rd.get("coleman_liau")
        if cl:
            _r("Readability", "Coleman-Liau", cl.cli_index, "Grade level")
        gf = rd.get("gunning_fog")
        if gf:
            _r("Readability", "Gunning Fog", gf.fog_index, "Grade level")
        dc = rd.get("dale_chall")
        if dc:
            _r("Readability", "Dale-Chall", dc.dale_chall_score, "Readability score")
        lw = rd.get("linsear")
        if lw:
            _r("Readability", "Linsear Write", lw.linsear_score, "Grade level")
        fc = rd.get("forcast")
        if fc:
            _r("Readability", "FORCAST", fc.forcast_score, "Grade level")
        pk = rd.get("psk")
        if pk:
            _r("Readability", "Powers-Sumner-Kearl", pk.psk_score, "Grade level")

    # -- Function words --
    fw = results.get("function_words")
    if fw:
        _r("Function Words", "Total Ratio", fw.total_function_word_ratio, "Function words / all tokens")
        _r("Function Words", "Diversity", fw.function_word_diversity, "Unique function words / total function words")

    # -- Character --
    ch = results.get("character")
    if ch:
        _r("Character", "Avg Word Length", ch.avg_word_length, "Characters per word")
        _r("Character", "Punctuation Density", ch.punctuation_density, "Punctuation per 100 words")
        _r("Character", "Vowel:Consonant", ch.vowel_consonant_ratio, "Vowel-to-consonant ratio")
        _r("Character", "Uppercase Ratio", ch.uppercase_ratio, "Uppercase letters / total letters")

    # -- Style markers --
    mk = results.get("markers")
    if mk:
        _r("Style", "Contraction Ratio", mk.contraction_ratio, "Contractions / (contractions + expanded)")
        _r("Style", "Intensifier Density", mk.intensifier_density, "Intensifiers per 100 words")
        _r("Style", "Hedge Density", mk.hedging_density, "Hedges per 100 words")
        _r("Style", "Modal Density", mk.modal_density, "Modal auxiliaries per 100 words")
        _r("Style", "Negation Density", mk.negation_density, "Negation markers per 100 words")
        _r("Style", "Exclamation Density", mk.exclamation_density, "Exclamation marks per 100 words")
        _r("Style", "Dash Density", mk.dash_density, "Dashes per 100 words")
        _r("Style", "Ellipsis Density", mk.ellipsis_density, "Ellipses per 100 words")

    # -- N-grams --
    ng = results.get("ngrams", {})
    cb = ng.get("char_bi")
    wb_ent = ng.get("word_bi")
    if cb:
        _r("N-grams", "Char Bigram Entropy", cb.entropy, "Character-level predictability")
    if wb_ent:
        _r("N-grams", "Word Bigram Entropy", wb_ent.entropy, "Word-level predictability")

    # -- Prosody --
    pr = results.get("prosody")
    if pr:
        _r("Prosody", "Syllables/Word", pr.mean_syllables_per_word, "Average syllables per word")
        _r("Prosody", "Rhythmic Regularity", pr.rhythmic_regularity, "Coefficient of variation (lower = more regular)")
        _r("Prosody", "Alliteration Density", pr.alliteration_density, "Alliterative pairs per 100 words")
        _r("Prosody", "Polysyllabic Ratio", pr.polysyllabic_ratio, "3+ syllable words / total")

    # -- Genre --
    gr = results.get("genre")
    if gr:
        _r("Genre", "Formality Score", gr.formality_score, "0-100 (higher = more formal)")
        _r("Genre", "Register", gr.register_classification, "Frozen / formal / consultative / casual / intimate")
        _r("Genre", "Predicted Genre", gr.predicted_genre, "Most likely genre classification")

    # -- Dialect --
    dl = results.get("dialect")
    if dl:
        _r("Dialect", "Detected Dialect", dl.dialect, "Dominant dialect")
        _r("Dialect", "British Score", dl.british_score, "British confidence")
        _r("Dialect", "American Score", dl.american_score, "American confidence")

    # -- Drift --
    dr = results.get("drift")
    if dr:
        _r("Drift", "Pattern", dr.pattern, "Detected consistency pattern")
        _r("Drift", "Confidence", dr.pattern_confidence, "Classification confidence")
        _r("Drift", "Mean Chi-Squared", dr.mean_chi_squared, "Average chi-squared across chunk pairs")
        _r("Drift", "Status", dr.status, "Analysis status")

    # -- Skipped --
    for tab_name, reason in results.get("_skipped", []):
        _r("Skipped", tab_name, "N/A", reason)


def _lexical_tab(ws: Any, lex: dict[str, Any]) -> None:
    """Write the Lexical Diversity detail tab."""
    if not lex:
        return

    def _row(name: str, val: Any, dist: Any = None) -> None:
        std = _fmt(dist.std) if dist else ""
        mn = _fmt(dist.min) if dist and hasattr(dist, "min") else ""
        mx = _fmt(dist.max) if dist and hasattr(dist, "max") else ""
        # Distribution objects use .range not .min/.max sometimes
        if not mn and dist:
            mn = _fmt(min(dist.values)) if dist.values else ""
        if not mx and dist:
            mx = _fmt(max(dist.values)) if dist.values else ""
        ws.append([name, _fmt(val), std, mn, mx])

    ttr = lex.get("ttr")
    if ttr:
        _row("TTR", ttr.ttr, ttr.ttr_dist)
        _row("Root TTR (Guiraud)", ttr.root_ttr, ttr.root_ttr_dist)
        _row("Log TTR (Herdan)", ttr.log_ttr, ttr.log_ttr_dist)
        _row("STTR", ttr.sttr, ttr.sttr_dist if ttr.metadata.get("sttr_available") else None)

    mtld = lex.get("mtld")
    if mtld:
        _row("MTLD (forward)", mtld.mtld_forward, mtld.mtld_forward_dist)
        _row("MTLD (backward)", mtld.mtld_backward, mtld.mtld_backward_dist)
        _row("MTLD (average)", mtld.mtld_average, mtld.mtld_average_dist)

    yule = lex.get("yule")
    if yule:
        _row("Yule's K", yule.yule_k, yule.yule_k_dist)
        _row("Yule's I", yule.yule_i, yule.yule_i_dist)

    hapax = lex.get("hapax")
    if hapax:
        _row("Hapax Ratio", hapax.hapax_ratio, hapax.hapax_ratio_dist)
        _row("Dis-Hapax Ratio", hapax.dis_hapax_ratio, hapax.dis_hapax_ratio_dist)
        _row("Sichel's S", hapax.sichel_s, hapax.sichel_s_dist)
        _row("Honore's R", hapax.honore_r, hapax.honore_r_dist)

    vocd = lex.get("vocd_d")
    if vocd:
        _row("VocD-D", vocd.d_parameter, vocd.d_parameter_dist)

    mattr = lex.get("mattr")
    if mattr:
        _row("MATTR", mattr.mattr_score, mattr.mattr_score_dist)

    hdd = lex.get("hdd")
    if hdd:
        _row("HDD", hdd.hdd_score, hdd.hdd_score_dist)

    msttr = lex.get("msttr")
    if msttr:
        _row("MSTTR", msttr.msttr_score, msttr.msttr_score_dist)


def _readability_tab(ws: Any, rd: dict[str, Any]) -> None:
    """Write the Readability detail tab."""
    def _row(name: str, score: Any, dist: Any = None) -> None:
        std = _fmt(dist.std) if dist else ""
        ws.append([name, _fmt(score), std])

    fl = rd.get("flesch")
    if fl:
        _row("Flesch Reading Ease", fl.reading_ease, fl.reading_ease_dist)
    sm = rd.get("smog")
    if sm:
        _row("SMOG Grade", sm.smog_index, sm.smog_index_dist)
    ar = rd.get("ari")
    if ar:
        _row("ARI", ar.ari_score, ar.ari_score_dist)
    cl = rd.get("coleman_liau")
    if cl:
        _row("Coleman-Liau", cl.cli_index, cl.cli_index_dist)
    gf = rd.get("gunning_fog")
    if gf:
        _row("Gunning Fog", gf.fog_index, gf.fog_index_dist)
    dc = rd.get("dale_chall")
    if dc:
        _row("Dale-Chall", dc.dale_chall_score, dc.dale_chall_score_dist)
        ws.append(["  Grade Level", dc.grade_level, ""])
    lw = rd.get("linsear")
    if lw:
        _row("Linsear Write", lw.linsear_score, lw.linsear_score_dist)
    fry = rd.get("fry")
    if fry:
        ws.append(["Fry Grade Level", fry.grade_level, ""])
        ws.append(["  Graph Zone", fry.graph_zone, ""])
    fc = rd.get("forcast")
    if fc:
        _row("FORCAST", fc.forcast_score, fc.forcast_score_dist)
    pk = rd.get("psk")
    if pk:
        _row("Powers-Sumner-Kearl", pk.psk_score, pk.psk_score_dist)


def _sentences_tab(ws: Any, syn: dict[str, Any]) -> None:
    """Write the Sentences detail tab."""
    ss = syn.get("sent_stats")
    if ss:
        ws.append(["Mean Sentence Length", _fmt(ss.mean_sentence_length), _fmt(ss.mean_sentence_length_dist.std)])
        ws.append(["Sentence Length Std Dev", _fmt(ss.sentence_length_std), ""])
        ws.append(["Min Sentence Length", _fmt(ss.min_sentence_length), ""])
        ws.append(["Max Sentence Length", _fmt(ss.max_sentence_length), ""])
        ws.append(["Sentence Count", ss.sentence_count, ""])

    st = syn.get("sent_types")
    if st:
        _section_row(ws, "Sentence Types")
        ws.append(["Declarative Ratio", _fmt(st.declarative_ratio), _fmt(st.declarative_ratio_dist.std)])
        ws.append(["Interrogative Ratio", _fmt(st.interrogative_ratio), _fmt(st.interrogative_ratio_dist.std)])
        ws.append(["Exclamatory Ratio", _fmt(st.exclamatory_ratio), _fmt(st.exclamatory_ratio_dist.std)])
        ws.append(["Imperative Ratio", _fmt(st.imperative_ratio), _fmt(st.imperative_ratio_dist.std)])

    adv = syn.get("advanced")
    if adv:
        _section_row(ws, "Advanced Syntactic")
        ws.append(["Mean Parse Tree Depth", _fmt(adv.mean_parse_tree_depth), ""])
        ws.append(["Mean T-Unit Length", _fmt(adv.mean_t_unit_length), ""])
        ws.append(["Clausal Density", _fmt(adv.clausal_density), ""])
        ws.append(["Dependent Clause Ratio", _fmt(adv.dependent_clause_ratio), ""])
        ws.append(["Passive Voice Ratio", _fmt(adv.passive_voice_ratio), ""])
        ws.append(["Subordination Index", _fmt(adv.subordination_index), ""])
        ws.append(["Coordination Index", _fmt(adv.coordination_index), ""])
        ws.append(["Complexity Score", _fmt(adv.sentence_complexity_score), ""])


def _function_words_tab(ws: Any, fw: Any) -> None:
    """Write the Function Words detail tab."""
    if not fw:
        return
    ws.append(["Total Function Word Ratio", _fmt(fw.total_function_word_ratio)])
    ws.append(["Function Word Diversity", _fmt(fw.function_word_diversity)])
    ws.append(["Determiner Ratio", _fmt(fw.determiner_ratio)])
    ws.append(["Preposition Ratio", _fmt(fw.preposition_ratio)])
    ws.append(["Conjunction Ratio", _fmt(fw.conjunction_ratio)])
    ws.append(["Pronoun Ratio", _fmt(fw.pronoun_ratio)])
    ws.append(["Auxiliary Ratio", _fmt(fw.auxiliary_ratio)])
    ws.append(["Particle Ratio", _fmt(fw.particle_ratio)])

    _section_row(ws, "Top 10 Most Frequent")
    for word, count in (fw.most_frequent_function_words or [])[:10]:
        ws.append([word, count])

    _section_row(ws, "Top 10 Least Frequent")
    for word, count in (fw.least_frequent_function_words or [])[:10]:
        ws.append([word, count])


def _style_markers_tab(ws: Any, mk: Any) -> None:
    """Write the Style Markers detail tab."""
    if not mk:
        return
    ws.append(["Contraction Ratio", _fmt(mk.contraction_ratio)])
    ws.append(["Contraction Count", mk.contraction_count])
    ws.append(["Expanded Form Count", mk.expanded_form_count])

    if mk.top_contractions:
        _section_row(ws, "Top Contractions")
        for word, count in mk.top_contractions[:10]:
            ws.append([word, count])

    _section_row(ws, "Densities (per 100 words)")
    ws.append(["Intensifier Density", _fmt(mk.intensifier_density)])
    ws.append(["Hedge Density", _fmt(mk.hedging_density)])
    ws.append(["Modal Density", _fmt(mk.modal_density)])
    ws.append(["Negation Density", _fmt(mk.negation_density)])
    ws.append(["Exclamation Density", _fmt(mk.exclamation_density)])
    ws.append(["Question Density", _fmt(mk.question_density)])
    ws.append(["Quotation Density", _fmt(mk.quotation_density)])
    ws.append(["Parenthetical Density", _fmt(mk.parenthetical_density)])
    ws.append(["Ellipsis Density", _fmt(mk.ellipsis_density)])
    ws.append(["Dash Density", _fmt(mk.dash_density)])
    ws.append(["Semicolon Density", _fmt(mk.semicolon_density)])
    ws.append(["Colon Density", _fmt(mk.colon_density)])

    _section_row(ws, "Modal Distribution")
    for modal, count in sorted(mk.modal_distribution.items(), key=lambda x: -x[1]):
        ws.append([modal, count])

    if mk.top_intensifiers:
        _section_row(ws, "Top Intensifiers")
        for word, count in mk.top_intensifiers[:10]:
            ws.append([word, count])

    if mk.top_hedges:
        _section_row(ws, "Top Hedges")
        for word, count in mk.top_hedges[:10]:
            ws.append([word, count])


def _character_tab(ws: Any, ch: Any) -> None:
    """Write the Character detail tab."""
    if not ch:
        return
    ws.append(["Avg Word Length", _fmt(ch.avg_word_length), _fmt(ch.avg_word_length_dist.std)])
    ws.append(["Avg Sentence Length (chars)", _fmt(ch.avg_sentence_length_chars), _fmt(ch.avg_sentence_length_chars_dist.std)])
    ws.append(["Punctuation Density", _fmt(ch.punctuation_density), _fmt(ch.punctuation_density_dist.std)])
    ws.append(["Punctuation Variety", _fmt(ch.punctuation_variety), ""])
    ws.append(["Vowel:Consonant Ratio", _fmt(ch.vowel_consonant_ratio), _fmt(ch.vowel_consonant_ratio_dist.std)])
    ws.append(["Digit Ratio", _fmt(ch.digit_ratio), _fmt(ch.digit_ratio_dist.std)])
    ws.append(["Uppercase Ratio", _fmt(ch.uppercase_ratio), _fmt(ch.uppercase_ratio_dist.std)])
    ws.append(["Whitespace Ratio", _fmt(ch.whitespace_ratio), _fmt(ch.whitespace_ratio_dist.std)])

    _section_row(ws, "Letter Frequencies (a-z)")
    for letter in sorted(ch.letter_frequency.keys()):
        ws.append([letter, _fmt(ch.letter_frequency[letter]), ""])


def _ngrams_tab(ws: Any, ng: dict[str, Any]) -> None:
    """Write the N-grams detail tab."""
    if not ng:
        return
    cb = ng.get("char_bi")
    wb_ent = ng.get("word_bi")
    if cb:
        ws.append(["Char Bigram Entropy", _fmt(cb.entropy)])
        ws.append(["Char Bigram Perplexity", _fmt(cb.perplexity)])
    if wb_ent:
        ws.append(["Word Bigram Entropy", _fmt(wb_ent.entropy)])
        ws.append(["Word Bigram Perplexity", _fmt(wb_ent.perplexity)])

    ext = ng.get("extended")
    if ext:
        if hasattr(ext, "top_word_trigrams") and ext.top_word_trigrams:
            _section_row(ws, "Top Word Trigrams")
            for ngram_str, freq in ext.top_word_trigrams[:15]:
                ws.append([ngram_str, freq])
        if hasattr(ext, "word_trigram_entropy"):
            ws.append(["Word Trigram Entropy", _fmt(ext.word_trigram_entropy)])
        if hasattr(ext, "top_word_4grams") and ext.top_word_4grams:
            _section_row(ws, "Top Word 4-grams")
            for ngram_str, freq in ext.top_word_4grams[:15]:
                ws.append([ngram_str, freq])
        if hasattr(ext, "word_4gram_entropy"):
            ws.append(["Word 4-gram Entropy", _fmt(ext.word_4gram_entropy)])


def _prosody_tab(ws: Any, pr: Any) -> None:
    """Write the Prosody detail tab."""
    if not pr:
        return
    ws.append(["Mean Syllables/Word", _fmt(pr.mean_syllables_per_word)])
    ws.append(["Syllable Std Dev", _fmt(pr.syllable_std_dev)])
    ws.append(["Polysyllabic Ratio (3+)", _fmt(pr.polysyllabic_ratio)])
    ws.append(["Monosyllabic Ratio", _fmt(pr.monosyllabic_ratio)])
    ws.append(["Rhythmic Regularity", _fmt(pr.rhythmic_regularity)])
    ws.append(["Alliteration Density", _fmt(pr.alliteration_density)])
    ws.append(["Assonance Density", _fmt(pr.assonance_density)])
    ws.append(["Consonance Density", _fmt(pr.consonance_density)])
    ws.append(["Sentence Rhythm Score", _fmt(pr.sentence_rhythm_score)])
    ws.append(["Avg Consonant Cluster", _fmt(pr.mean_consonant_cluster_length)])

    # Syllable distribution: Total and Unique per bucket (1–7, 8+)
    meta = pr.metadata if hasattr(pr, "metadata") else {}
    syl_total = meta.get("syllable_total_by_bucket", {})
    syl_unique = meta.get("syllable_unique_by_bucket", {})
    if syl_total:
        _section_row(ws, "Syllable Distribution")
        # Sub-header
        row_num = ws.max_row + 1
        ws.cell(row=row_num, column=1, value="Syllables")
        ws.cell(row=row_num, column=2, value="Total")
        ws.cell(row=row_num, column=3, value="Unique")
        for col in range(1, 4):
            c = ws.cell(row=row_num, column=col)
            c.font = Font(bold=True)
        # Data rows
        buckets = [str(b) for b in range(1, 8)] + ["8+"]
        for bucket in buckets:
            ws.append([bucket, syl_total.get(bucket, 0), syl_unique.get(bucket, 0)])


def _cohesion_tab(ws: Any, co: Any) -> None:
    """Write the Cohesion detail tab."""
    if not co:
        return
    ws.append(["Pronoun Density", _fmt(co.pronoun_density)])
    ws.append(["Demonstrative Density", _fmt(co.demonstrative_density)])
    ws.append(["Connective Density", _fmt(co.connective_density)])
    ws.append(["Additive Connective Ratio", _fmt(co.additive_connective_ratio)])
    ws.append(["Adversative Connective Ratio", _fmt(co.adversative_connective_ratio)])
    ws.append(["Causal Connective Ratio", _fmt(co.causal_connective_ratio)])
    ws.append(["Temporal Connective Ratio", _fmt(co.temporal_connective_ratio)])
    ws.append(["Adjacent Sentence Overlap", _fmt(co.adjacent_sentence_overlap)])
    ws.append(["Word Repetition Ratio", _fmt(co.word_repetition_ratio)])
    ws.append(["Lexical Chain Count", co.lexical_chain_count])
    ws.append(["Mean Chain Length", _fmt(co.mean_chain_length)])


def _genre_tab(ws: Any, gr: Any) -> None:
    """Write the Genre & Register detail tab."""
    if not gr:
        return
    ws.append(["Formality Score", _fmt(gr.formality_score)])
    ws.append(["Register Classification", gr.register_classification])
    ws.append(["Predicted Genre", gr.predicted_genre])
    ws.append(["Latinate Ratio", _fmt(gr.latinate_ratio)])
    ws.append(["Nominalization Density", _fmt(gr.nominalization_density)])
    ws.append(["Passive Voice Density", _fmt(gr.passive_voice_density)])
    ws.append(["First Person Ratio", _fmt(gr.first_person_ratio)])
    ws.append(["Second Person Ratio", _fmt(gr.second_person_ratio)])
    ws.append(["Third Person Ratio", _fmt(gr.third_person_ratio)])
    ws.append(["Narrative Marker Density", _fmt(gr.narrative_marker_density)])
    ws.append(["Dialogue Ratio", _fmt(gr.dialogue_ratio)])


def _dialect_tab(ws: Any, dl: Any) -> None:
    """Write the Dialect detail tab."""
    if not dl:
        return
    ws.append(["Detected Dialect", dl.dialect])
    ws.append(["Confidence", _fmt(dl.confidence)])
    ws.append(["British Score", _fmt(dl.british_score)])
    ws.append(["American Score", _fmt(dl.american_score)])
    ws.append(["Markedness Score", _fmt(dl.markedness_score)])
    ws.append(["Eye Dialect Count", dl.eye_dialect_count])
    ws.append(["Eye Dialect Ratio", _fmt(dl.eye_dialect_ratio)])

    # Marker breakdown by level
    if dl.markers_by_level:
        _section_row(ws, "Markers by Level")
        for level, markers in dl.markers_by_level.items():
            if markers:
                for marker, count in sorted(markers.items(), key=lambda x: -x[1])[:10]:
                    ws.append([f"{level}: {marker}", count])


def _drift_tab(ws: Any, dr: Any) -> None:
    """Write the Style Drift detail tab."""
    if not dr:
        return
    ws.append(["Status", dr.status])
    ws.append(["Pattern", dr.pattern])
    ws.append(["Pattern Confidence", _fmt(dr.pattern_confidence)])
    ws.append(["Mean Chi-Squared", _fmt(dr.mean_chi_squared)])
    ws.append(["Std Chi-Squared", _fmt(dr.std_chi_squared)])
    ws.append(["Min Chi-Squared", _fmt(dr.min_chi_squared)])
    ws.append(["Max Chi-Squared", _fmt(dr.max_chi_squared)])
    ws.append(["Trend", _fmt(dr.trend)])
    ws.append(["Window Count", dr.window_count])

    # Per-pair chi-squared values
    if hasattr(dr, "pairwise_scores") and dr.pairwise_scores:
        _section_row(ws, "Pairwise Chi-Squared")
        for i, score in enumerate(dr.pairwise_scores):
            if isinstance(score, dict):
                chi_sq = score.get("chi_squared", "")
                ws.append([f"Pair {i}-{i + 1}", _fmt(chi_sq)])
            else:
                ws.append([f"Pair {i}-{i + 1}", _fmt(score)])


def _repetition_tab(ws: Any, rep_uni: Any, rep_ng: Any) -> None:
    """Write the Repetition detail tab."""
    if rep_uni and hasattr(rep_uni, "repetitive_words") and rep_uni.repetitive_words:
        for rw in rep_uni.repetitive_words[:50]:
            ws.append([rw.word, rw.count, _fmt(rw.expected_count), _fmt(rw.repetition_score)])

    if rep_ng and hasattr(rep_ng, "repetitive_ngrams") and rep_ng.repetitive_ngrams:
        _section_row(ws, "Repetitive N-grams")
        for rn in rep_ng.repetitive_ngrams[:50]:
            ws.append([" ".join(rn.ngram), rn.count, "", _fmt(rn.frequency_per_10k)])


def _bnc_detail_tabs(wb: Workbook, bnc: Any) -> None:
    """Write the 3 BNC detail tabs (overused, underused, not-in-bnc)."""
    try:
        from pystylometry.lexical.word_class import classify_word
    except ImportError:
        classify_word = None  # type: ignore[assignment]

    # -- Overused --
    ws = wb.create_sheet("BNC Overused")
    ws.append(["Word", "Observed", "Expected", "Ratio", "WordNet", "GNgram", "Classification"])
    _style_header(ws)
    for w in sorted(bnc.overused, key=lambda x: x.ratio or 0, reverse=True):
        cls = classify_word(w.word).label if classify_word else ""
        ws.append([
            w.word, w.observed, _fmt(w.expected), _fmt(w.ratio),
            "true" if w.in_wordnet else "false",
            "true" if w.in_gngram else "false",
            cls,
        ])
    _auto_width(ws)
    _align_all(ws)
    _set_row_heights(ws)

    # -- Underused --
    ws = wb.create_sheet("BNC Underused")
    ws.append(["Word", "Observed", "Expected", "Ratio", "WordNet", "GNgram", "Classification"])
    _style_header(ws)
    for w in sorted(bnc.underused, key=lambda x: x.ratio or 0):
        cls = classify_word(w.word).label if classify_word else ""
        ws.append([
            w.word, w.observed, _fmt(w.expected), _fmt(w.ratio),
            "true" if w.in_wordnet else "false",
            "true" if w.in_gngram else "false",
            cls,
        ])
    _auto_width(ws)
    _align_all(ws)
    _set_row_heights(ws)

    # -- Not in BNC --
    ws = wb.create_sheet("BNC Not Found")
    ws.append(["Word", "Observed", "WordNet", "GNgram", "Classification"])
    _style_header(ws)
    for w in sorted(bnc.not_in_bnc, key=lambda x: x.observed or 0, reverse=True):
        cls = classify_word(w.word).label if classify_word else ""
        ws.append([
            w.word, w.observed,
            "true" if w.in_wordnet else "false",
            "true" if w.in_gngram else "false",
            cls,
        ])
    _auto_width(ws)
    _align_all(ws)
    _set_row_heights(ws)
