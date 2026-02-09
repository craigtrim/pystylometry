"""Tests for BNC meta-analysis (comparative summary across authors).

Tests the read/write roundtrip for the bnc_meta module, verifying:

- Single-file summary extraction
- Multi-file directory reading with alphabetical sort
- Excel output structure (sheet names, headers, data cells)
- Robustness to missing summary sheets and lock files

Related GitHub Issue:
    #55 -- Add bnc-meta CLI for comparative BNC summary analysis across authors
    https://github.com/craigtrim/pystylometry/issues/55
"""

import pytest
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

from pystylometry.lexical.bnc_meta import (
    AuthorSummary,
    METRIC_LABELS,
    PCT_METRIC_LABELS,
    read_author_summary,
    read_bnc_summaries,
    write_bnc_meta_excel,
)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def _create_bnc_excel(path: Path, stats: dict[str, tuple[str, str]]) -> None:
    """Create a minimal BNC Excel file with a summary sheet.

    ``stats`` maps metric label -> (count_str, pct_str).
    """
    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("summary")

    # Top-10 placeholder rows (rows 1-12)
    ws.cell(row=1, column=1, value="Top 10 Overused")

    # Stats header at row 14
    ws.cell(row=14, column=1, value="Metric")
    ws.cell(row=14, column=2, value="Count")
    ws.cell(row=14, column=3, value="% of Total Tokens")

    # Stats rows 15-19
    labels = ["Total Tokens", "Unique Tokens", "Overused", "Underused", "Not in BNC"]
    for i, label in enumerate(labels):
        row = 15 + i
        count, pct = stats.get(label, ("0", ""))
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=count)
        ws.cell(row=row, column=3, value=pct)

    wb.save(path)


_SAMPLE_STATS = {
    "Total Tokens": ("3,709,051", ""),
    "Unique Tokens": ("69,051", "1.86%"),
    "Overused": ("18,246", "0.49%"),
    "Underused": ("15,092", "0.41%"),
    "Not in BNC": ("31,158", "0.84%"),
}

_SAMPLE_STATS_B = {
    "Total Tokens": ("2,500,000", ""),
    "Unique Tokens": ("45,000", "1.80%"),
    "Overused": ("12,000", "0.48%"),
    "Underused": ("10,000", "0.40%"),
    "Not in BNC": ("20,000", "0.80%"),
}


# ===================================================================
# read_author_summary
# ===================================================================


class TestReadAuthorSummary:
    """Extract summary from a single BNC Excel file."""

    def test_reads_all_fields(self, tmp_path: Path) -> None:
        xlsx = tmp_path / "Ray Bradbury.xlsx"
        _create_bnc_excel(xlsx, _SAMPLE_STATS)

        result = read_author_summary(xlsx)
        assert result is not None
        assert result.author == "Ray Bradbury"
        assert result.total_tokens == 3709051
        assert result.unique_tokens == 69051
        assert result.overused == 18246
        assert result.underused == 15092
        assert result.not_in_bnc == 31158
        assert result.pct_unique_tokens == pytest.approx(1.86)
        assert result.pct_overused == pytest.approx(0.49)
        assert result.pct_underused == pytest.approx(0.41)
        assert result.pct_not_in_bnc == pytest.approx(0.84)

    def test_returns_none_for_missing_summary_sheet(self, tmp_path: Path) -> None:
        xlsx = tmp_path / "bad.xlsx"
        wb = Workbook()
        wb.save(xlsx)

        result = read_author_summary(xlsx)
        assert result is None

    def test_returns_none_for_missing_metric_header(self, tmp_path: Path) -> None:
        xlsx = tmp_path / "no_header.xlsx"
        wb = Workbook()
        wb.remove(wb.active)
        ws = wb.create_sheet("summary")
        ws.cell(row=1, column=1, value="something else")
        wb.save(xlsx)

        result = read_author_summary(xlsx)
        assert result is None


# ===================================================================
# read_bnc_summaries
# ===================================================================


class TestReadBncSummaries:
    """Read multiple BNC files from a directory."""

    def test_reads_multiple_sorted_alphabetically(self, tmp_path: Path) -> None:
        _create_bnc_excel(tmp_path / "Zebra Author.xlsx", _SAMPLE_STATS)
        _create_bnc_excel(tmp_path / "Alpha Author.xlsx", _SAMPLE_STATS_B)
        _create_bnc_excel(tmp_path / "Middle Author.xlsx", _SAMPLE_STATS)

        summaries = read_bnc_summaries(tmp_path)
        assert len(summaries) == 3
        assert summaries[0].author == "Alpha Author"
        assert summaries[1].author == "Middle Author"
        assert summaries[2].author == "Zebra Author"

    def test_skips_lock_files(self, tmp_path: Path) -> None:
        _create_bnc_excel(tmp_path / "Author.xlsx", _SAMPLE_STATS)
        # Create a lock file
        (tmp_path / "~$Author.xlsx").write_bytes(b"lock")

        summaries = read_bnc_summaries(tmp_path)
        assert len(summaries) == 1

    def test_skips_files_without_summary(self, tmp_path: Path) -> None:
        _create_bnc_excel(tmp_path / "Good.xlsx", _SAMPLE_STATS)
        # Bad file without summary sheet
        wb = Workbook()
        wb.save(tmp_path / "Bad.xlsx")

        summaries = read_bnc_summaries(tmp_path)
        assert len(summaries) == 1
        assert summaries[0].author == "Good"

    def test_empty_dir_returns_empty_list(self, tmp_path: Path) -> None:
        summaries = read_bnc_summaries(tmp_path)
        assert summaries == []


# ===================================================================
# write_bnc_meta_excel
# ===================================================================


class TestWriteBncMetaExcel:
    """Write comparative workbook and verify structure."""

    @pytest.fixture()
    def two_author_output(self, tmp_path: Path) -> Path:
        summaries = [
            AuthorSummary(
                author="Author A",
                total_tokens=1000,
                unique_tokens=100,
                overused=20,
                underused=10,
                not_in_bnc=30,
                pct_unique_tokens=10.00,
                pct_overused=2.00,
                pct_underused=1.00,
                pct_not_in_bnc=3.00,
            ),
            AuthorSummary(
                author="Author B",
                total_tokens=2000,
                unique_tokens=200,
                overused=40,
                underused=20,
                not_in_bnc=60,
                pct_unique_tokens=10.00,
                pct_overused=2.00,
                pct_underused=1.00,
                pct_not_in_bnc=3.00,
            ),
        ]
        out = tmp_path / "comparison.xlsx"
        write_bnc_meta_excel(summaries, out)
        return out

    def test_output_file_exists(self, two_author_output: Path) -> None:
        assert two_author_output.exists()

    def test_sheet_names(self, two_author_output: Path) -> None:
        wb = load_workbook(two_author_output, read_only=True)
        assert wb.sheetnames == ["counts", "percentages"]
        wb.close()

    def test_counts_sheet_dimensions(self, two_author_output: Path) -> None:
        wb = load_workbook(two_author_output, read_only=True)
        ws = wb["counts"]
        rows = list(ws.rows)
        # 1 header + 2 author rows = 3 rows
        assert len(rows) == 3
        # 1 author col + 5 metric cols = 6 columns
        assert len(rows[0]) == 6
        wb.close()

    def test_percentages_sheet_dimensions(self, two_author_output: Path) -> None:
        wb = load_workbook(two_author_output, read_only=True)
        ws = wb["percentages"]
        rows = list(ws.rows)
        # 1 header + 2 author rows = 3 rows
        assert len(rows) == 3
        # 1 author col + 4 metric cols = 5 columns
        assert len(rows[0]) == 5
        wb.close()

    def test_counts_header_row(self, two_author_output: Path) -> None:
        wb = load_workbook(two_author_output, read_only=True)
        ws = wb["counts"]
        header = [c.value for c in list(ws.rows)[0]]
        assert header == ["Author", "Total Tokens", "Unique Tokens", "Overused", "Underused", "Not in BNC"]
        wb.close()

    def test_counts_data_values(self, two_author_output: Path) -> None:
        wb = load_workbook(two_author_output, read_only=True)
        ws = wb["counts"]
        # Row 2 = Author A
        assert ws.cell(row=2, column=1).value == "Author A"
        assert ws.cell(row=2, column=2).value == 1000
        # Row 3 = Author B
        assert ws.cell(row=3, column=1).value == "Author B"
        assert ws.cell(row=3, column=2).value == 2000
        wb.close()

    def test_counts_number_format(self, two_author_output: Path) -> None:
        wb = load_workbook(two_author_output)
        ws = wb["counts"]
        assert ws.cell(row=2, column=2).number_format == "#,##0"
        wb.close()

    def test_percentages_data_values(self, two_author_output: Path) -> None:
        wb = load_workbook(two_author_output, read_only=True)
        ws = wb["percentages"]
        # Row 2 = Author A
        assert ws.cell(row=2, column=1).value == "Author A"
        assert ws.cell(row=2, column=2).value == pytest.approx(10.00)
        # Row 3 = Author B
        assert ws.cell(row=3, column=1).value == "Author B"
        assert ws.cell(row=3, column=2).value == pytest.approx(10.00)
        wb.close()

    def test_percentages_number_format(self, two_author_output: Path) -> None:
        wb = load_workbook(two_author_output)
        ws = wb["percentages"]
        assert ws.cell(row=2, column=2).number_format == "0.00\\%"
        wb.close()


# ===================================================================
# Roundtrip: read -> write -> verify
# ===================================================================


class TestRoundtrip:
    """End-to-end: create BNC files, read, write meta, verify."""

    def test_full_roundtrip(self, tmp_path: Path) -> None:
        # Create source Excel files
        src = tmp_path / "source"
        src.mkdir()
        _create_bnc_excel(src / "Author A.xlsx", _SAMPLE_STATS)
        _create_bnc_excel(src / "Author B.xlsx", _SAMPLE_STATS_B)

        # Read
        summaries = read_bnc_summaries(src)
        assert len(summaries) == 2

        # Write
        out = tmp_path / "meta.xlsx"
        write_bnc_meta_excel(summaries, out)

        # Verify
        wb = load_workbook(out, read_only=True)
        ws_counts = wb["counts"]
        ws_pct = wb["percentages"]

        # Row 2 = Author A (sorted alphabetically)
        assert ws_counts.cell(row=2, column=1).value == "Author A"
        assert ws_counts.cell(row=2, column=2).value == 3709051
        # Row 3 = Author B
        assert ws_counts.cell(row=3, column=1).value == "Author B"
        assert ws_counts.cell(row=3, column=2).value == 2500000

        # Percentages: Author A unique tokens %
        assert ws_pct.cell(row=2, column=2).value == pytest.approx(1.86)
        # Author B unique tokens %
        assert ws_pct.cell(row=3, column=2).value == pytest.approx(1.80)

        wb.close()
