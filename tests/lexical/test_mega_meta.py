"""Tests for mega meta-analysis (comparative N-gram summary across authors).

Tests the read/write roundtrip for the mega_meta module, verifying:

- Single-file N-gram extraction
- Multi-file directory reading (including subdirectories) with alphabetical sort
- Excel output structure (sheet name, headers, data cells, number formats)
- Robustness to missing N-grams sheets and lock files
"""

import pytest
from pathlib import Path

from openpyxl import Workbook, load_workbook

from pystylometry.lexical.mega_meta import (
    NgramSummary,
    NGRAM_METRIC_LABELS,
    read_ngram_summary,
    read_ngram_summaries,
    write_mega_meta_excel,
)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def _create_mega_excel(path: Path, metrics: dict[str, float]) -> None:
    """Create a minimal mega Excel file with an N-grams sheet."""
    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("N-grams")

    ws.cell(row=1, column=1, value="Metric")
    ws.cell(row=1, column=2, value="Value")

    labels = [
        "Char Bigram Entropy",
        "Char Bigram Perplexity",
        "Word Bigram Entropy",
        "Word Bigram Perplexity",
    ]
    for i, label in enumerate(labels):
        ws.cell(row=2 + i, column=1, value=label)
        ws.cell(row=2 + i, column=2, value=metrics.get(label, 0.0))

    # Extra rows that should be ignored
    ws.cell(row=6, column=1, value="Top Word Trigrams")
    ws.cell(row=7, column=1, value="foo bar baz")
    ws.cell(row=7, column=2, value=42)

    wb.save(path)


_SAMPLE_A = {
    "Char Bigram Entropy": 7.7006,
    "Char Bigram Perplexity": 208.5789,
    "Word Bigram Entropy": 9.7204,
    "Word Bigram Perplexity": 851.0615,
}

_SAMPLE_B = {
    "Char Bigram Entropy": 7.6655,
    "Char Bigram Perplexity": 204.0521,
    "Word Bigram Entropy": 9.4706,
    "Word Bigram Perplexity": 742.7981,
}


# ===================================================================
# read_ngram_summary
# ===================================================================


class TestReadNgramSummary:
    """Extract N-gram stats from a single mega Excel file."""

    def test_reads_all_fields(self, tmp_path: Path) -> None:
        xlsx = tmp_path / "G K Chesterton.xlsx"
        _create_mega_excel(xlsx, _SAMPLE_A)

        result = read_ngram_summary(xlsx)
        assert result is not None
        assert result.author == "G K Chesterton"
        assert result.char_bigram_entropy == pytest.approx(7.7006)
        assert result.char_bigram_perplexity == pytest.approx(208.5789)
        assert result.word_bigram_entropy == pytest.approx(9.7204)
        assert result.word_bigram_perplexity == 851

    def test_returns_none_for_missing_ngrams_sheet(self, tmp_path: Path) -> None:
        xlsx = tmp_path / "bad.xlsx"
        wb = Workbook()
        wb.save(xlsx)

        result = read_ngram_summary(xlsx)
        assert result is None

    def test_returns_none_for_incomplete_metrics(self, tmp_path: Path) -> None:
        xlsx = tmp_path / "partial.xlsx"
        wb = Workbook()
        wb.remove(wb.active)
        ws = wb.create_sheet("N-grams")
        ws.cell(row=1, column=1, value="Metric")
        ws.cell(row=1, column=2, value="Value")
        ws.cell(row=2, column=1, value="Char Bigram Entropy")
        ws.cell(row=2, column=2, value=7.5)
        # Missing other 3 metrics
        wb.save(xlsx)

        result = read_ngram_summary(xlsx)
        assert result is None


# ===================================================================
# read_ngram_summaries
# ===================================================================


class TestReadNgramSummaries:
    """Read multiple mega files from a directory tree."""

    def test_reads_flat_directory(self, tmp_path: Path) -> None:
        _create_mega_excel(tmp_path / "Zebra.xlsx", _SAMPLE_A)
        _create_mega_excel(tmp_path / "Alpha.xlsx", _SAMPLE_B)

        summaries = read_ngram_summaries(tmp_path)
        assert len(summaries) == 2
        assert summaries[0].author == "Alpha"
        assert summaries[1].author == "Zebra"

    def test_reads_subdirectories(self, tmp_path: Path) -> None:
        decade = tmp_path / "1920s"
        decade.mkdir()
        _create_mega_excel(decade / "Author.xlsx", _SAMPLE_A)

        summaries = read_ngram_summaries(tmp_path)
        assert len(summaries) == 1
        assert summaries[0].author == "Author"

    def test_skips_lock_files(self, tmp_path: Path) -> None:
        _create_mega_excel(tmp_path / "Author.xlsx", _SAMPLE_A)
        (tmp_path / "~$Author.xlsx").write_bytes(b"lock")

        summaries = read_ngram_summaries(tmp_path)
        assert len(summaries) == 1

    def test_skips_files_without_ngrams(self, tmp_path: Path) -> None:
        _create_mega_excel(tmp_path / "Good.xlsx", _SAMPLE_A)
        wb = Workbook()
        wb.save(tmp_path / "Bad.xlsx")

        summaries = read_ngram_summaries(tmp_path)
        assert len(summaries) == 1
        assert summaries[0].author == "Good"

    def test_empty_dir_returns_empty_list(self, tmp_path: Path) -> None:
        summaries = read_ngram_summaries(tmp_path)
        assert summaries == []


# ===================================================================
# write_mega_meta_excel
# ===================================================================


class TestWriteMegaMetaExcel:
    """Write comparative workbook and verify structure."""

    @pytest.fixture()
    def two_author_output(self, tmp_path: Path) -> Path:
        summaries = [
            NgramSummary(
                author="Author A",
                char_bigram_entropy=7.7006,
                char_bigram_perplexity=208.5789,
                word_bigram_entropy=9.7204,
                word_bigram_perplexity=851,
            ),
            NgramSummary(
                author="Author B",
                char_bigram_entropy=7.6655,
                char_bigram_perplexity=204.0521,
                word_bigram_entropy=9.4706,
                word_bigram_perplexity=743,
            ),
        ]
        out = tmp_path / "comparison.xlsx"
        write_mega_meta_excel(summaries, out)
        return out

    def test_output_file_exists(self, two_author_output: Path) -> None:
        assert two_author_output.exists()

    def test_sheet_names(self, two_author_output: Path) -> None:
        wb = load_workbook(two_author_output, read_only=True)
        assert wb.sheetnames == ["N-grams"]
        wb.close()

    def test_dimensions(self, two_author_output: Path) -> None:
        wb = load_workbook(two_author_output, read_only=True)
        ws = wb["N-grams"]
        rows = list(ws.rows)
        # 1 header + 2 author rows = 3
        assert len(rows) == 3
        # 1 author col + 4 metric cols = 5
        assert len(rows[0]) == 5
        wb.close()

    def test_header_row(self, two_author_output: Path) -> None:
        wb = load_workbook(two_author_output, read_only=True)
        ws = wb["N-grams"]
        header = [c.value for c in list(ws.rows)[0]]
        assert header == ["Author"] + NGRAM_METRIC_LABELS
        wb.close()

    def test_data_values(self, two_author_output: Path) -> None:
        wb = load_workbook(two_author_output, read_only=True)
        ws = wb["N-grams"]
        assert ws.cell(row=2, column=1).value == "Author A"
        assert ws.cell(row=2, column=2).value == pytest.approx(7.7006)
        assert ws.cell(row=2, column=5).value == 851
        assert ws.cell(row=3, column=1).value == "Author B"
        assert ws.cell(row=3, column=2).value == pytest.approx(7.6655)
        wb.close()

    def test_number_format(self, two_author_output: Path) -> None:
        wb = load_workbook(two_author_output)
        ws = wb["N-grams"]
        assert ws.cell(row=2, column=2).number_format == "#,##0.0000"
        assert ws.cell(row=2, column=5).number_format == "#,##0"
        wb.close()


# ===================================================================
# Roundtrip: read -> write -> verify
# ===================================================================


class TestRoundtrip:
    """End-to-end: create mega files, read, write meta, verify."""

    def test_full_roundtrip(self, tmp_path: Path) -> None:
        src = tmp_path / "source"
        src.mkdir()
        _create_mega_excel(src / "Author A.xlsx", _SAMPLE_A)
        _create_mega_excel(src / "Author B.xlsx", _SAMPLE_B)

        summaries = read_ngram_summaries(src)
        assert len(summaries) == 2

        out = tmp_path / "meta.xlsx"
        write_mega_meta_excel(summaries, out)

        wb = load_workbook(out, read_only=True)
        ws = wb["N-grams"]

        assert ws.cell(row=2, column=1).value == "Author A"
        assert ws.cell(row=2, column=2).value == pytest.approx(7.7006)
        assert ws.cell(row=3, column=1).value == "Author B"
        assert ws.cell(row=3, column=2).value == pytest.approx(7.6655)
        assert ws.cell(row=3, column=5).value == 743

        wb.close()

    def test_roundtrip_with_subdirectories(self, tmp_path: Path) -> None:
        src = tmp_path / "source"
        d1 = src / "1920s"
        d2 = src / "1950s"
        d1.mkdir(parents=True)
        d2.mkdir(parents=True)
        _create_mega_excel(d1 / "Author A.xlsx", _SAMPLE_A)
        _create_mega_excel(d2 / "Author B.xlsx", _SAMPLE_B)

        summaries = read_ngram_summaries(src)
        assert len(summaries) == 2

        out = tmp_path / "meta.xlsx"
        write_mega_meta_excel(summaries, out)

        wb = load_workbook(out, read_only=True)
        ws = wb["N-grams"]
        assert ws.cell(row=2, column=1).value == "Author A"
        assert ws.cell(row=3, column=1).value == "Author B"
        wb.close()
